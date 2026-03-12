"""
FTIR Data Processing Tools
==========================
Module 1: Thickness calculation (Reflectance Method)
Module 2: Glass density and refractive index calculation
Module 3: H2O peak baseline correction and fitting
Module 4: CO3²⁻ carbonate doublet fitting (replaces CO2_bestfit.xls Solver)
Module 5: Concentration calculations (H2O wt%, CO2 ppm) via Beer-Lambert law

Pipeline functions:
  process_sample() — full single-sample pipeline (Modules 1-5)
  process_batch()  — batch processing with CSV/Excel output

Replicates the logic from:
  - Reflectance Method Template.xlsx
  - Glass density and index of refraction Template.xls
  - CO2_bestfit.xls
  - Master Table Template.xlsx
"""

import json
import os
import statistics
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union

import numpy as np
from scipy.interpolate import interp1d
from scipy.optimize import minimize

# ---------------------------------------------------------------------------
# Module 1 — Thickness Calculation (Reflectance Method)
# ---------------------------------------------------------------------------
# Formula from Reflectance Method Template.xlsx:
#   thickness (µm) = N / (2 * n * (ν_high - ν_low)) * 10000
#
# Where:
#   N        = number of interference fringes (peaks)
#   n        = refractive index of the material
#   ν_high   = higher wavenumber (cm⁻¹)
#   ν_low    = lower wavenumber (cm⁻¹)
#   * 10000  = convert from cm to µm
# ---------------------------------------------------------------------------

# Pre-loaded refractive indices from the template
REFRACTIVE_INDICES = {
    "quartz":   1.54,
    "sanidine": 1.518,   # range: 1.518–1.531
    "olivine":  1.683,   # Fo=0.85
}

# Olivine refractive index: linear interpolation between endmembers
# Forsterite (Fo100): n_alpha=1.635, n_beta=1.651, n_gamma=1.670 -> mean~1.652
# Fayalite   (Fo0):   n_alpha=1.827, n_beta=1.869, n_gamma=1.879 -> mean~1.858
# Using mean of 3 principal indices (consistent with FTIR transmission geometry).
# Reference: Deer, Howie & Zussman (1992), An Introduction to Rock-Forming Minerals.
_OLIVINE_N_FO100 = 1.652   # forsterite endmember mean n
_OLIVINE_N_FO0   = 1.858   # fayalite endmember mean n


def olivine_refractive_index(fo: float) -> float:
    """Calculate olivine refractive index from forsterite content.

    Linear interpolation between forsterite (Fo100, n=1.652) and
    fayalite (Fo0, n=1.858) endmembers.

    Parameters
    ----------
    fo : float
        Forsterite content as mole fraction (0-1) or mol% (0-100).
        Values > 1 are interpreted as mol%.

    Returns
    -------
    float : mean refractive index
    """
    if fo > 1.0:
        fo = fo / 100.0
    if not (0.0 <= fo <= 1.0):
        raise ValueError(f"Fo must be between 0 and 1 (or 0-100), got {fo}")
    return _OLIVINE_N_FO0 + fo * (_OLIVINE_N_FO100 - _OLIVINE_N_FO0)


def calculate_thickness(
    wavenumber_high: float,
    wavenumber_low: float,
    num_fringes: int,
    refractive_index: float,
) -> Dict[str, float]:
    """Calculate sample thickness from FTIR reflectance interference fringes.

    Parameters
    ----------
    wavenumber_high : float
        Higher wavenumber bound (cm⁻¹).
    wavenumber_low : float
        Lower wavenumber bound (cm⁻¹).
    num_fringes : int
        Number of interference fringes counted between the two wavenumbers.
    refractive_index : float
        Refractive index of the material.

    Returns
    -------
    dict with keys:
        "thickness_um"  – thickness in micrometers (µm)
        "thickness_cm"  – thickness in centimeters (cm)
    """
    if wavenumber_high <= wavenumber_low:
        raise ValueError("wavenumber_high must be greater than wavenumber_low")
    if num_fringes <= 0:
        raise ValueError("num_fringes must be positive")
    if refractive_index <= 0:
        raise ValueError("refractive_index must be positive")

    delta_nu = wavenumber_high - wavenumber_low
    thickness_cm = num_fringes / (2.0 * refractive_index * delta_nu)
    thickness_um = thickness_cm * 10000.0

    return {
        "thickness_um": thickness_um,
        "thickness_cm": thickness_cm,
    }


def calculate_thickness_multi(
    measurements: List[Tuple[float, float, int]],
    refractive_index: float,
) -> Dict[str, float]:
    """Calculate average thickness from multiple measurements.

    Parameters
    ----------
    measurements : list of (wavenumber_high, wavenumber_low, num_fringes)
    refractive_index : float

    Returns
    -------
    dict with keys:
        "thicknesses_um"  – list of individual thickness values (µm)
        "average_um"      – average thickness (µm)
        "stdev_um"        – standard deviation (µm)
        "average_cm"      – average thickness (cm)
    """
    thicknesses = []
    for wn_high, wn_low, n_fringes in measurements:
        result = calculate_thickness(wn_high, wn_low, n_fringes, refractive_index)
        thicknesses.append(result["thickness_um"])

    avg = statistics.mean(thicknesses)
    std = statistics.stdev(thicknesses) if len(thicknesses) > 1 else 0.0

    return {
        "thicknesses_um": thicknesses,
        "average_um": avg,
        "stdev_um": std,
        "average_cm": avg / 10000.0,
    }


def _best_fringe_run(extrema_wn: np.ndarray, n_refr: float,
                     tol: float = 0.20) -> Tuple[int, int]:
    """Find the longest consecutive run of consistent fringe intervals.

    For each pair of adjacent extrema, compute per-interval thickness.
    Then find the longest consecutive sub-sequence where every value is
    within *tol* (relative) of the sub-sequence median.

    Parameters
    ----------
    extrema_wn : array
        Wavenumber positions of detected extrema (sorted).
    n_refr : float
        Refractive index.
    tol : float
        Relative tolerance (default 0.20 = 20%).

    Returns
    -------
    (start, end) : indices into *extrema_wn* defining the best run
        (inclusive on both ends).  The run contains (end - start) intervals
        from (end - start + 1) extrema.
    """
    n_pts = len(extrema_wn)
    if n_pts < 3:
        return (0, n_pts - 1)

    # Per-interval thicknesses
    t_intervals = []
    for i in range(n_pts - 1):
        dv = abs(extrema_wn[i + 1] - extrema_wn[i])
        if dv > 0:
            t_intervals.append(1.0 / (2.0 * n_refr * dv) * 10000.0)
        else:
            t_intervals.append(np.inf)

    # Find longest consecutive run where all values are within tol of
    # the run's median.  Use expanding-window approach:
    # for each start, extend end as far as possible.
    best_start, best_end = 0, 1  # at least 2 extrema = 1 interval
    best_len = 1

    for s in range(len(t_intervals)):
        for e in range(s + 1, len(t_intervals) + 1):
            run = t_intervals[s:e]
            med = statistics.median(run)
            if med == 0 or med == np.inf:
                break
            if all(abs(v - med) / med <= tol for v in run):
                if e - s > best_len:
                    best_len = e - s
                    best_start = s
                    best_end = s + e - s  # = e
            else:
                break  # further extension will only get worse

    # Convert interval indices back to extrema indices
    # intervals [best_start, best_end) use extrema [best_start, best_end]
    return (best_start, best_start + best_len)


def calculate_thickness_from_spectrum(
    reflectance_csv: Union[str, Path],
    refractive_index: float,
    wn_range: Tuple[float, float] = (1800.0, 3000.0),
    sg_window: int = 31,
    sg_order: int = 3,
    min_prominence: float = 0.003,
    consistency_tol: float = 0.20,
    save_figure: bool = True,
    figure_path: Optional[Union[str, Path]] = None,
) -> Dict:
    """Calculate thickness from reflectance spectrum by automatic fringe detection.

    Reads a reflectance spectrum, applies Savitzky-Golay smoothing to reduce
    noise, detects interference fringe extrema (peaks and troughs), then
    selects the **best sub-region** (longest consecutive run of consistent
    fringe spacings) and calculates thickness from that region only.

    Algorithm:
      1. Savitzky-Golay filter to smooth spectrum
      2. Detect maxima and minima independently
      3. For each set of extrema, compute per-interval thickness and find
         the longest consecutive run where all intervals agree within
         *consistency_tol* of the run median
      4. Use only the selected sub-region for thickness calculation
      5. Report average +/- uncertainty from per-interval statistics

    Parameters
    ----------
    reflectance_csv : str or Path
        Path to reflectance spectrum file (wavenumber vs reflectance).
    refractive_index : float
        Refractive index of the host mineral (e.g. olivine).
        Use olivine_refractive_index(Fo) to calculate from Fo content.
    wn_range : (float, float)
        Wavenumber range (cm-1) for fringe analysis. Default (1800, 3000).
        Should be a clean region free of absorption bands.
    sg_window : int
        Savitzky-Golay filter window length (must be odd). Default 31.
    sg_order : int
        Savitzky-Golay polynomial order. Default 3.
    min_prominence : float
        Minimum prominence for peak detection. Default 0.003.
    consistency_tol : float
        Relative tolerance for selecting consistent fringe intervals.
        Default 0.20 (20%).  Intervals whose thickness deviates more than
        this from the run median are excluded.
    save_figure : bool
        If True, save a diagnostic plot. Default True.
    figure_path : str or Path or None
        Explicit path for the figure. If None, saved next to the CSV.

    Returns
    -------
    dict with keys:
        "average_um"       -- average thickness (um)
        "stdev_um"         -- uncertainty (um)
        "average_cm"       -- average thickness (cm)
        "thickness_maxima" -- thickness from selected maxima sub-region (um)
        "thickness_minima" -- thickness from selected minima sub-region (um)
        "n_maxima"         -- number of maxima detected (total)
        "n_minima"         -- number of minima detected (total)
        "n_maxima_used"    -- number of maxima in selected sub-region
        "n_minima_used"    -- number of minima in selected sub-region
        "maxima_wn"        -- wavenumbers of all detected maxima (cm-1)
        "minima_wn"        -- wavenumbers of all detected minima (cm-1)
        "best_maxima_wn"   -- wavenumbers of selected maxima (cm-1)
        "best_minima_wn"   -- wavenumbers of selected minima (cm-1)
        "refractive_index" -- refractive index used
        "figure_path"      -- path to saved figure (or None)
    """
    from scipy.signal import savgol_filter, find_peaks

    # Read spectrum
    wn_full, ref_full = _read_spectrum_csv(reflectance_csv)

    # Extract analysis range
    lo, hi = min(wn_range), max(wn_range)
    mask = (wn_full >= lo) & (wn_full <= hi)
    if mask.sum() < 50:
        raise ValueError(
            f"Only {mask.sum()} points in range [{lo}, {hi}] cm-1. "
            "Check spectrum coverage.")
    wn = wn_full[mask]
    ref = ref_full[mask]

    # Estimate data spacing
    data_spacing = abs(float(np.mean(np.diff(wn))))  # cm-1 per point

    # Savitzky-Golay smoothing
    if sg_window % 2 == 0:
        sg_window += 1  # must be odd
    ref_smooth = savgol_filter(ref, sg_window, sg_order)

    # Estimate minimum distance between fringes (conservative: 40% of
    # expected spacing for a ~30 um sample)
    rough_thickness_cm = 0.003  # 30 um initial guess
    expected_spacing = 1.0 / (2.0 * refractive_index * rough_thickness_cm)
    min_dist_pts = max(5, int(expected_spacing * 0.4 / data_spacing))

    peaks_idx, _ = find_peaks(ref_smooth,
                              distance=min_dist_pts,
                              prominence=min_prominence)
    troughs_idx, _ = find_peaks(-ref_smooth,
                                distance=min_dist_pts,
                                prominence=min_prominence)

    maxima_wn = wn[peaks_idx]
    minima_wn = wn[troughs_idx]
    n_max = len(maxima_wn)
    n_min = len(minima_wn)

    # --- Select best sub-region for maxima and minima independently ---
    t_max = None
    best_max_wn = np.array([])
    n_max_used = 0
    if n_max >= 2:
        s, e = _best_fringe_run(maxima_wn, refractive_index, tol=consistency_tol)
        best_max_wn = maxima_wn[s:e + 1]
        n_max_used = len(best_max_wn)
        if n_max_used >= 2:
            N_int = n_max_used - 1
            dnu = abs(best_max_wn[-1] - best_max_wn[0])
            if dnu > 0:
                t_max = N_int / (2.0 * refractive_index * dnu) * 10000.0

    t_min = None
    best_min_wn = np.array([])
    n_min_used = 0
    if n_min >= 2:
        s, e = _best_fringe_run(minima_wn, refractive_index, tol=consistency_tol)
        best_min_wn = minima_wn[s:e + 1]
        n_min_used = len(best_min_wn)
        if n_min_used >= 2:
            N_int = n_min_used - 1
            dnu = abs(best_min_wn[-1] - best_min_wn[0])
            if dnu > 0:
                t_min = N_int / (2.0 * refractive_index * dnu) * 10000.0

    # Combine results
    thicknesses = [t for t in [t_max, t_min] if t is not None]
    if not thicknesses:
        raise ValueError(
            f"Could not detect enough fringes: {n_max} maxima, {n_min} minima. "
            "Try adjusting wn_range, sg_window, or min_prominence.")

    avg_um = statistics.mean(thicknesses)

    # Uncertainty from per-interval thicknesses within selected sub-regions
    per_interval = []
    if n_max_used >= 3:
        for i in range(n_max_used - 1):
            dv = abs(best_max_wn[i + 1] - best_max_wn[i])
            if dv > 0:
                per_interval.append(
                    1.0 / (2.0 * refractive_index * dv) * 10000.0)
    if n_min_used >= 3:
        for i in range(n_min_used - 1):
            dv = abs(best_min_wn[i + 1] - best_min_wn[i])
            if dv > 0:
                per_interval.append(
                    1.0 / (2.0 * refractive_index * dv) * 10000.0)
    if len(per_interval) >= 2:
        std_um = statistics.stdev(per_interval)
    elif len(thicknesses) == 2:
        std_um = abs(thicknesses[0] - thicknesses[1]) / 2.0
    else:
        std_um = 0.0

    # Diagnostic plot
    fig_saved = None
    if save_figure:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt

        fig, ax = plt.subplots(figsize=(12, 5))
        ax.plot(wn, ref, "k-", lw=0.5, alpha=0.4, label="Raw reflectance")
        ax.plot(wn, ref_smooth, "b-", lw=1.0, alpha=0.6,
                label="Savitzky-Golay smoothed")
        # All detected extrema (grey, smaller)
        if n_max > 0:
            ax.plot(maxima_wn, ref_smooth[peaks_idx], "v",
                    color="grey", ms=6, alpha=0.5,
                    label=f"All maxima ({n_max})")
        if n_min > 0:
            ax.plot(minima_wn, ref_smooth[troughs_idx], "^",
                    color="grey", ms=6, alpha=0.5,
                    label=f"All minima ({n_min})")
        # Selected sub-region extrema (bold colours)
        if len(best_max_wn) > 0:
            _sel_idx = [i for i, w in enumerate(maxima_wn)
                        if w in best_max_wn]
            ax.plot(best_max_wn, ref_smooth[peaks_idx[_sel_idx]], "rv", ms=10,
                    label=f"Selected maxima ({n_max_used})")
        if len(best_min_wn) > 0:
            _sel_idx = [i for i, w in enumerate(minima_wn)
                        if w in best_min_wn]
            ax.plot(best_min_wn, ref_smooth[troughs_idx[_sel_idx]], "g^",
                    ms=10, label=f"Selected minima ({n_min_used})")
        ax.set_xlabel("Wavenumber (cm$^{-1}$)")
        ax.set_ylabel("Reflectance")
        t_str = f"{avg_um:.1f} +/- {std_um:.1f} um"
        ax.set_title(
            f"Interference Fringes - {Path(reflectance_csv).stem}\n"
            f"Thickness = {t_str}  |  n = {refractive_index:.4f}")
        ax.legend(fontsize=8, loc="best")
        ax.invert_xaxis()
        ax.grid(True, alpha=0.3)
        plt.tight_layout()

        if figure_path is None:
            figure_path = Path(reflectance_csv).with_name(
                Path(reflectance_csv).stem + "_fringes.png")
        fig_saved = str(figure_path)
        fig.savefig(fig_saved, dpi=150)
        plt.close(fig)

    return {
        "average_um": avg_um,
        "stdev_um": std_um,
        "average_cm": avg_um / 10000.0,
        "thickness_maxima": t_max,
        "thickness_minima": t_min,
        "n_maxima": n_max,
        "n_minima": n_min,
        "n_maxima_used": n_max_used,
        "n_minima_used": n_min_used,
        "maxima_wn": maxima_wn.tolist(),
        "minima_wn": minima_wn.tolist(),
        "best_maxima_wn": best_max_wn.tolist(),
        "best_minima_wn": best_min_wn.tolist(),
        "refractive_index": refractive_index,
        "figure_path": fig_saved,
    }


# ---------------------------------------------------------------------------
# Module 2 — Glass Density and Refractive Index
# ---------------------------------------------------------------------------
# Density: based on partial molar volumes (Lange, 1999; Luhr, 2001)
#   density = Σ(Xi * MWi) / Σ(Xi * V̄i)
#
# Refractive index: Church & Johnson (1980)
#   n = 1 + Σ(wt_fraction_i * coefficient_i)
# ---------------------------------------------------------------------------

# Molecular weights used in the Excel template (back-calculated from data)
_MOLECULAR_WEIGHTS = {
    "SiO2":  60.09,
    "TiO2":  79.90,
    "Al2O3": 101.96,
    "Fe2O3": 159.70,
    "FeO":   71.85,
    "MgO":   40.31,
    "CaO":   56.08,
    "Na2O":  61.80,
    "K2O":   94.20,
    "H2O":   18.02,
}

# Partial molar volumes at 25°C, 1 bar (cm³/mol)
# Source: R. Lange pers. commun. 1999, as cited in Luhr (2001)
_PARTIAL_MOLAR_VOLUMES = {
    "SiO2":  27.01,
    "TiO2":  13.205,
    "Al2O3": 37.76,
    "Fe2O3": 29.63125,
    "FeO":   10.50,
    "MgO":    8.81,
    "CaO":   13.03,
    "Na2O":  19.88,
    "K2O":   33.63,
    "H2O":   13.93,
}

# Church & Johnson (1980) coefficients for (n-1)
_REFRACTIVE_INDEX_COEFFICIENTS = {
    "SiO2":  0.46,
    "TiO2":  1.158,
    "Al2O3": 0.581,
    "Fe2O3": 1.09,
    "FeO":   0.897,
    "MgO":   0.903,
    "CaO":   0.795,
    "Na2O":  0.505,
    "K2O":   0.495,
    # H2O has no coefficient in the template
}


def _split_iron(
    composition: Dict[str, float],
    fe3_over_fetotal: float = 0.15,
) -> Dict[str, float]:
    """Split FeOT into Fe2O3 and FeO if needed.

    If composition has "FeOT" (or "FeO_total"), it is split into "Fe2O3" and
    "FeO" using the given Fe³⁺/Fe_total ratio.  If both "Fe2O3" and "FeO"
    are already present, the composition is returned unchanged.
    """
    comp = dict(composition)

    feot_key = None
    for key in ("FeOT", "FeO_total", "FeOt"):
        if key in comp:
            feot_key = key
            break

    if feot_key is not None:
        feot = comp.pop(feot_key)
        # Fe2O3 = FeOT * ratio * (MW_Fe2O3 / (2 * MW_FeO))
        comp["Fe2O3"] = feot * fe3_over_fetotal * (_MOLECULAR_WEIGHTS["Fe2O3"]
                                                    / (2 * _MOLECULAR_WEIGHTS["FeO"]))
        comp["FeO"] = feot * (1.0 - fe3_over_fetotal)

    return comp


def calculate_density(
    composition: Dict[str, float],
    fe3_over_fetotal: float = 0.15,
) -> Dict[str, float]:
    """Calculate glass density from oxide weight-percent composition.

    Parameters
    ----------
    composition : dict
        Oxide wt% values.  Recognised keys:
        SiO2, TiO2, Al2O3, Fe2O3, FeO, MgO, CaO, Na2O, K2O, H2O.
        Alternatively, provide "FeOT" instead of separate Fe2O3/FeO.
    fe3_over_fetotal : float
        Fe³⁺ / total-Fe ratio (default 0.15, as in the template).

    Returns
    -------
    dict with keys:
        "density_gcc"   – density in g/cm³
        "density_kgm3"  – density in kg/m³
        "molar_volume"  – molar volume (cm³/mol)
        "gfw"           – gram formula weight on mole-fraction basis
        "moles"         – dict of moles per oxide
        "mole_fractions" – dict of mole fractions
    """
    comp = _split_iron(composition, fe3_over_fetotal)

    # Step 1: wt% → moles
    moles = {}
    for oxide, wt in comp.items():
        if oxide in _MOLECULAR_WEIGHTS and wt > 0:
            moles[oxide] = wt / _MOLECULAR_WEIGHTS[oxide]

    total_moles = sum(moles.values())
    if total_moles == 0:
        raise ValueError("Total moles is zero — check composition input")

    # Step 2: mole fractions
    mole_fracs = {ox: m / total_moles for ox, m in moles.items()}

    # Step 3: molar volume = Σ(Xi * V̄i)
    molar_volume = sum(
        mole_fracs[ox] * _PARTIAL_MOLAR_VOLUMES[ox]
        for ox in mole_fracs
        if ox in _PARTIAL_MOLAR_VOLUMES
    )

    # Step 4: gram formula weight = Σ(Xi * MWi)
    gfw = sum(
        mole_fracs[ox] * _MOLECULAR_WEIGHTS[ox]
        for ox in mole_fracs
    )

    # Step 5: density
    density_gcc = gfw / molar_volume
    density_kgm3 = density_gcc * 1000.0

    return {
        "density_gcc": density_gcc,
        "density_kgm3": density_kgm3,
        "molar_volume": molar_volume,
        "gfw": gfw,
        "moles": moles,
        "mole_fractions": mole_fracs,
    }


def get_refractive_index(
    composition: Dict[str, float],
    fe3_over_fetotal: float = 0.15,
) -> Dict[str, float]:
    """Calculate refractive index using Church & Johnson (1980).

    Parameters
    ----------
    composition : dict
        Oxide wt% values (same format as calculate_density).
    fe3_over_fetotal : float
        Fe³⁺ / total-Fe ratio (default 0.15).

    Returns
    -------
    dict with keys:
        "n"             – refractive index
        "n_minus_1"     – the (n-1) value
        "contributions" – per-oxide contributions to (n-1)
    """
    comp = _split_iron(composition, fe3_over_fetotal)

    # Convert wt% to weight fractions
    total_wt = sum(comp.values())
    contributions = {}

    for oxide, coeff in _REFRACTIVE_INDEX_COEFFICIENTS.items():
        wt = comp.get(oxide, 0.0)
        wt_frac = wt / 100.0  # wt% to fraction (template uses wt%/100)
        contributions[oxide] = wt_frac * coeff

    n_minus_1 = sum(contributions.values())
    n = 1.0 + n_minus_1

    return {
        "n": n,
        "n_minus_1": n_minus_1,
        "contributions": contributions,
    }


# ---------------------------------------------------------------------------
# Module 3 — H2O Peak Baseline Correction
# ---------------------------------------------------------------------------
# Replaces manual baseline correction in Spectragryph.
#
# Approach:
#   1. Read spectrum CSV (wavenumber vs absorbance)
#   2. Extract the region around the O-H stretch peak (~3550 cm⁻¹)
#   3. Fit a linear baseline through anchor points at both edges of the
#      peak window (averaging a small margin to reduce noise)
#   4. Subtract the baseline → baseline-corrected absorbance
#   5. Find peak height and peak position
# ---------------------------------------------------------------------------

# Supported CSV delimiters (auto-detected)
_CSV_DELIMITERS = [",", "\t", ";", " "]


def _read_spectrum_csv(filepath: Union[str, Path]) -> Tuple[np.ndarray, np.ndarray]:
    """Read a two-column spectrum file (wavenumber, absorbance).

    Handles common export formats from Spectragryph, OPUS, Omnic, etc.:
      - Comma, tab, semicolon, or space delimited
      - Optional header row(s)
      - Wavenumber in ascending or descending order

    Returns
    -------
    wavenumber : ndarray, sorted ascending
    absorbance : ndarray, matching order
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Spectrum file not found: {filepath}")

    text = filepath.read_text(encoding="utf-8-sig")  # handles BOM
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    # Auto-detect delimiter: try each, pick the one that gives 2 columns
    delimiter = None
    for delim in _CSV_DELIMITERS:
        parts = lines[-1].split(delim)
        nums = [p for p in parts if p.strip()]
        if len(nums) >= 2:
            try:
                float(nums[0])
                float(nums[1])
                delimiter = delim
                break
            except ValueError:
                continue
    if delimiter is None:
        raise ValueError(f"Cannot detect delimiter in {filepath.name}")

    wavenumbers = []
    absorbances = []
    for line in lines:
        parts = [p.strip() for p in line.split(delimiter) if p.strip()]
        if len(parts) < 2:
            continue
        try:
            wn = float(parts[0])
            ab = float(parts[1])
            wavenumbers.append(wn)
            absorbances.append(ab)
        except ValueError:
            continue  # skip header rows

    if len(wavenumbers) == 0:
        raise ValueError(f"No numeric data found in {filepath.name}")

    wn = np.array(wavenumbers)
    ab = np.array(absorbances)

    # Ensure ascending wavenumber order
    if wn[0] > wn[-1]:
        wn = wn[::-1]
        ab = ab[::-1]

    return wn, ab


def fit_h2o_peak(
    spectrum_csv_path: Union[str, Path],
    peak_range: Tuple[float, float] = (2200, 3800),
    baseline_low_range: Tuple[float, float] = (2200, 2400),
    baseline_high_range: Optional[Tuple[float, float]] = None,
    peak_target_wn: float = 3550.0,
    anchor_width: int = 10,
    compute_uncertainty: bool = False,
    save_figure: bool = True,
    figure_path: Optional[Union[str, Path]] = None,
) -> Dict:
    """Baseline-correct the O-H stretch peak and measure its height.

    The baseline is a linear fit through two anchor regions:
      - Low-wavenumber anchor: 2200-2400 cm-1 (flat, absorption-free region)
      - High-wavenumber anchor: 3700-3800 cm-1 (above the O-H peak)

    Parameters
    ----------
    spectrum_csv_path : str or Path
        Path to the spectrum CSV file (wavenumber vs absorbance).
    peak_range : (float, float)
        Full wavenumber window to extract (cm-1).
        Default (2200, 3800) covers the baseline reference + O-H peak.
    baseline_low_range : (float, float)
        Wavenumber range for the low-side baseline anchor (cm-1).
        Default (2200, 2400) -- a flat region free of absorption features.
    baseline_high_range : (float, float) or None
        Wavenumber range for the high-side baseline anchor.
        If None, defaults to (peak_range[1] - 100, peak_range[1]),
        i.e. the top 100 cm-1 of the peak window.
    peak_target_wn : float
        Wavenumber at which to read the baseline-corrected absorbance
        (cm-1).  Default 3550.0 (O-H stretch peak position).
        The value is linearly interpolated from surrounding data points.
    anchor_width : int
        Number of data points averaged at each anchor region to define
        baseline anchor points (reduces noise sensitivity).  Default 10.
    compute_uncertainty : bool
        If True, estimate baseline uncertainty by sliding the low anchor
        through sub-windows within baseline_low_range.  Returns additional
        keys "peak_height_std" and "peak_height_range".  Default False.
    save_figure : bool
        If True, save a diagnostic plot (default True).
    figure_path : str or Path or None
        Explicit path for the figure.  If None, saved next to the CSV
        with suffix ``_h2o_baseline.png``.

    Returns
    -------
    dict with keys:
        "peak_height"      -- baseline-corrected absorbance at peak_target_wn
        "peak_wavenumber"  -- the target wavenumber used (cm-1)
        "baseline_corrected_absorbance" -- ndarray of corrected values
        "wavenumber"       -- ndarray of wavenumbers in the full region
        "baseline"         -- ndarray of the fitted linear baseline
        "raw_absorbance"   -- ndarray of original absorbance in region
        "figure_path"      -- path to saved figure (or None)
    If compute_uncertainty is True, also includes:
        "peak_height_std"   -- std of peak heights from anchor variation
        "peak_height_range" -- (min, max) of peak heights from anchor variation
    """
    wn_full, ab_full = _read_spectrum_csv(spectrum_csv_path)

    lo, hi = min(peak_range), max(peak_range)

    # Extract the full region (includes baseline reference zone + peak)
    mask = (wn_full >= lo) & (wn_full <= hi)
    if mask.sum() < 20:
        raise ValueError(
            f"Not enough data points in range [{lo}, {hi}] cm-1. "
            f"Found {mask.sum()} points."
        )

    wn = wn_full[mask]
    ab = ab_full[mask]

    # --- Low-side anchor (default: 2200-2400 cm-1, flat reference region) ---
    bl_lo = min(baseline_low_range)
    bl_hi = max(baseline_low_range)
    lo_mask = (wn >= bl_lo) & (wn <= bl_hi)
    if lo_mask.sum() < 3:
        raise ValueError(
            f"Not enough points in baseline_low_range [{bl_lo}, {bl_hi}]. "
            f"Found {lo_mask.sum()} points."
        )
    wn_lo_anchor = wn[lo_mask].mean()
    ab_lo_anchor = ab[lo_mask].mean()

    # --- High-side anchor (default: top 100 cm-1 of peak_range) ---
    if baseline_high_range is None:
        baseline_high_range = (hi - 100, hi)
    bh_lo = min(baseline_high_range)
    bh_hi = max(baseline_high_range)
    hi_mask = (wn >= bh_lo) & (wn <= bh_hi)
    if hi_mask.sum() < 3:
        raise ValueError(
            f"Not enough points in baseline_high_range [{bh_lo}, {bh_hi}]. "
            f"Found {hi_mask.sum()} points."
        )
    wn_hi_anchor = wn[hi_mask].mean()
    ab_hi_anchor = ab[hi_mask].mean()

    # --- Linear baseline through the two anchors ---
    slope = (ab_hi_anchor - ab_lo_anchor) / (wn_hi_anchor - wn_lo_anchor)
    intercept = ab_lo_anchor - slope * wn_lo_anchor
    baseline = slope * wn + intercept

    # Baseline-corrected absorbance
    corrected = ab - baseline

    # Peak height at target wavenumber (interpolated)
    peak_wn = float(peak_target_wn)
    sort_idx = np.argsort(wn)
    peak_height = float(np.interp(peak_target_wn, wn[sort_idx], corrected[sort_idx]))

    # --- Anchor variation uncertainty ---
    anchor_unc = {}
    if compute_uncertainty:
        lo_range_width = bl_hi - bl_lo  # e.g. 200 cm-1
        sub_width = lo_range_width / 3.0  # ~67 cm-1 sub-windows
        step = sub_width / 2.0  # 50% overlap
        peak_heights_var = []
        pos = bl_lo
        while pos + sub_width <= bl_hi + 0.1:
            sub_mask = (wn >= pos) & (wn <= pos + sub_width)
            if sub_mask.sum() >= 3:
                wn_sub = wn[sub_mask].mean()
                ab_sub = ab[sub_mask].mean()
                sl = (ab_hi_anchor - ab_sub) / (wn_hi_anchor - wn_sub)
                ic = ab_sub - sl * wn_sub
                corr_sub = ab - (sl * wn + ic)
                ph = float(np.interp(peak_target_wn, wn[sort_idx],
                                     corr_sub[sort_idx]))
                peak_heights_var.append(ph)
            pos += step
        if len(peak_heights_var) >= 2:
            anchor_unc["peak_height_std"] = float(np.std(peak_heights_var))
            anchor_unc["peak_height_range"] = (
                float(min(peak_heights_var)),
                float(max(peak_heights_var)),
            )
        else:
            anchor_unc["peak_height_std"] = 0.0
            anchor_unc["peak_height_range"] = (peak_height, peak_height)

    # Plotting
    fig_saved = None
    if save_figure:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt

        fig, axes = plt.subplots(2, 1, figsize=(12, 9), sharex=True)

        # --- Top panel: raw + baseline + anchor regions ---
        ax1 = axes[0]
        ax1.plot(wn, ab, "k-", linewidth=0.8, label="Raw spectrum")
        ax1.plot(wn, baseline, "r--", linewidth=1.0, label="Linear baseline")
        ax1.plot(
            [wn_lo_anchor, wn_hi_anchor],
            [ab_lo_anchor, ab_hi_anchor],
            "ro", markersize=7, zorder=5, label="Anchor points",
        )
        # Shade anchor regions
        ax1.axvspan(bl_lo, bl_hi, color="green", alpha=0.12,
                     label=f"Low anchor ({bl_lo:.0f}-{bl_hi:.0f})")
        ax1.axvspan(bh_lo, bh_hi, color="orange", alpha=0.12,
                     label=f"High anchor ({bh_lo:.0f}-{bh_hi:.0f})")
        ax1.set_ylabel("Absorbance")
        ax1.set_title(
            f"H$_2$O Peak Baseline Correction — {Path(spectrum_csv_path).stem}"
        )
        ax1.legend(fontsize=9)
        ax1.set_xlim(hi + 50, lo - 50)
        ax1.grid(True, alpha=0.3)

        # --- Bottom panel: corrected + peak annotation ---
        ax2 = axes[1]
        ax2.fill_between(wn, 0, corrected, where=(corrected > 0),
                         color="steelblue", alpha=0.3)
        ax2.plot(wn, corrected, "b-", linewidth=0.8, label="Baseline-corrected")
        ax2.axhline(0, color="gray", linewidth=0.5, linestyle="--")

        ax2.annotate(
            f"Peak: {peak_height:.4f}\n@ {peak_wn:.1f} cm$^{{-1}}$",
            xy=(peak_wn, peak_height),
            xytext=(peak_wn + (hi - lo) * 0.08, peak_height * 0.80),
            fontsize=10,
            arrowprops=dict(arrowstyle="->", color="red", lw=1.2),
            bbox=dict(boxstyle="round,pad=0.3", fc="lightyellow", ec="gray"),
        )
        ax2.plot(peak_wn, peak_height, "rv", markersize=8)

        ax2.set_xlabel("Wavenumber (cm$^{-1}$)")
        ax2.set_ylabel("Corrected Absorbance")
        ax2.legend(fontsize=9)
        ax2.set_xlim(hi + 50, lo - 50)
        ax2.grid(True, alpha=0.3)

        plt.tight_layout()

        if figure_path is None:
            figure_path = Path(spectrum_csv_path).with_name(
                Path(spectrum_csv_path).stem + "_h2o_baseline.png"
            )
        fig_saved = str(figure_path)
        fig.savefig(fig_saved, dpi=150)
        plt.close(fig)

    result = {
        "peak_height": float(peak_height),
        "peak_wavenumber": float(peak_wn),
        "baseline_corrected_absorbance": corrected,
        "wavenumber": wn,
        "baseline": baseline,
        "raw_absorbance": ab,
        "figure_path": fig_saved,
    }
    result.update(anchor_unc)
    return result


def batch_fit_h2o(
    csv_folder_path: Union[str, Path],
    peak_range: Tuple[float, float] = (2200, 3800),
    baseline_low_range: Tuple[float, float] = (2200, 2400),
    baseline_high_range: Optional[Tuple[float, float]] = None,
    anchor_width: int = 10,
    save_figures: bool = True,
    csv_extensions: Tuple[str, ...] = (".csv", ".CSV", ".txt", ".TXT", ".dat", ".DAT"),
) -> List[Dict]:
    """Batch-process all spectrum CSVs in a folder.

    Parameters
    ----------
    csv_folder_path : str or Path
        Directory containing spectrum CSV files.
    peak_range : (float, float)
        Full wavenumber window passed to fit_h2o_peak.
    baseline_low_range : (float, float)
        Low-side baseline anchor range passed to fit_h2o_peak.
    baseline_high_range : (float, float) or None
        High-side baseline anchor range passed to fit_h2o_peak.
    anchor_width : int
        Passed to fit_h2o_peak.
    save_figures : bool
        Whether to save per-file diagnostic plots.
    csv_extensions : tuple of str
        File extensions to include.

    Returns
    -------
    list of dicts, each containing:
        "sample"         – filename (stem)
        "peak_height"    – baseline-corrected peak absorbance
        "peak_wavenumber" – wavenumber at peak (cm⁻¹)
        "file"           – full file path
        "figure_path"    – path to saved figure (or None)
    """
    folder = Path(csv_folder_path)
    if not folder.is_dir():
        raise NotADirectoryError(f"Not a directory: {folder}")

    files = sorted(
        f for f in folder.iterdir()
        if f.is_file() and f.suffix in csv_extensions
    )

    if not files:
        raise FileNotFoundError(
            f"No spectrum files ({', '.join(csv_extensions)}) found in {folder}"
        )

    results = []
    for fpath in files:
        try:
            res = fit_h2o_peak(
                fpath,
                peak_range=peak_range,
                baseline_low_range=baseline_low_range,
                baseline_high_range=baseline_high_range,
                anchor_width=anchor_width,
                save_figure=save_figures,
            )
            results.append({
                "sample": fpath.stem,
                "peak_height": res["peak_height"],
                "peak_wavenumber": res["peak_wavenumber"],
                "file": str(fpath),
                "figure_path": res["figure_path"],
            })
        except Exception as e:
            print(f"  WARNING: Skipping {fpath.name}: {e}")

    # Print summary table
    print(f"\n{'Sample':<30s} {'Peak Height':>12s} {'Peak WN (cm-1)':>15s}")
    print("-" * 60)
    for r in results:
        print(f"{r['sample']:<30s} {r['peak_height']:>12.4f} {r['peak_wavenumber']:>15.1f}")

    return results


# ---------------------------------------------------------------------------
# Module 4 — CO3²⁻ Carbonate Doublet Fitting (replaces CO2_bestfit.xls)
# ---------------------------------------------------------------------------
# Model from CO2_bestfit.xls 'CO32-calc' sheet:
#
#   fitted(ν) = c_offset + c_bkg * E(ν) + c_1630 * F(ν) + c_CO3 * G(ν) + c_st * H(ν)
#
# Where:
#   E(ν) = standard background spectrum (WOK162DV.US3)
#   F(ν) = 1630 cm⁻¹ H₂O bending mode standard (normalized, peak=1 at 1635)
#   G(ν) = CO₃²⁻ doublet standard (~1430 + ~1515 cm⁻¹, normalized at 1527)
#   H(ν) = straight line component (index 1, 2, 3, ...)
#
# 5 coefficients optimized by Solver (least-squares):
#   c_offset  – constant offset
#   c_bkg     – background scaling
#   c_1630    – 1630 H₂O peak amplitude
#   c_CO3     – CO₃²⁻ doublet amplitude
#   c_st      – linear slope correction
#
# Objective: minimize Σ(fitted - data)² over the fitting range (~1350–1800 cm⁻¹)
#
# Outputs:
#   - CO₃²⁻ absorbance by subtraction: T(ν) = data - (fit without CO₃ term)
#   - Peak height of T at ~1515 cm⁻¹ and ~1430 cm⁻¹
#   - R² goodness of fit
# ---------------------------------------------------------------------------

def _load_reference_spectra() -> Dict[str, np.ndarray]:
    """Load CO₃²⁻ reference spectra from the bundled JSON file.

    The JSON was extracted from CO2_bestfit.xls and contains 259 data points
    covering 2198–1203 cm⁻¹ with ~3.86 cm⁻¹ spacing.

    Returns dict with keys: wavenumber, E_bkg, F_1630, G_CO3
    """
    json_path = Path(__file__).parent / "co2_reference_spectra.json"
    if not json_path.exists():
        raise FileNotFoundError(
            f"Reference spectra file not found: {json_path}\n"
            "Run the extraction script or ensure co2_reference_spectra.json "
            "is in the same directory as ftir_tools.py."
        )
    with open(json_path, "r") as f:
        data = json.load(f)
    return {k: np.array(v) for k, v in data.items()}


def _load_pca_components(wn_ref: np.ndarray) -> Dict[str, np.ndarray]:
    """Load PyIRoGlass PCA baseline and H2Om,1635 components.

    Interpolates the 596-point PyIRoGlass grid (1252-2399 cm-1)
    onto the provided reference wavenumber grid.

    Source: Shi et al. (2024) PyIRoGlass, doi:10.30909/vol.07.02.471501
    Data: BaselineAvgPC.npz + H2Om1635PC.npz from the PyIRoGlass GitHub.

    Returns dict with keys (all interpolated onto wn_ref):
        baseline_mean, baseline_pc1..pc4,
        h2om_mean, h2om_pc1, h2om_pc2
    """
    json_path = Path(__file__).parent / "pyiroglass_pca_components.json"
    if not json_path.exists():
        raise FileNotFoundError(
            f"PyIRoGlass PCA components not found: {json_path}\n"
            "Download from the PyIRoGlass GitHub repository."
        )
    with open(json_path, "r") as f:
        pca = json.load(f)
    wn_pca = np.array(pca["wavenumber"])

    # Interpolate each component onto the target grid
    # wn_pca is ascending (1252->2399), wn_ref may be descending
    sort_pca = np.argsort(wn_pca)
    wn_asc = wn_pca[sort_pca]

    result = {}
    for key_json, key_out in [
        ("baseline_mean", "baseline_mean"),
        ("baseline_pc1", "baseline_pc1"),
        ("baseline_pc2", "baseline_pc2"),
        ("baseline_pc3", "baseline_pc3"),
        ("baseline_pc4", "baseline_pc4"),
        ("h2om_1635_mean", "h2om_mean"),
        ("h2om_1635_pc1", "h2om_pc1"),
        ("h2om_1635_pc2", "h2om_pc2"),
    ]:
        vals = np.array(pca[key_json])[sort_pca]
        result[key_out] = np.interp(wn_ref, wn_asc, vals)
    return result


def _interpolate_to_reference_grid(
    wn_sample: np.ndarray,
    ab_sample: np.ndarray,
    wn_ref: np.ndarray,
) -> np.ndarray:
    """Interpolate sample spectrum onto the reference wavenumber grid.

    Uses linear interpolation.  Reference wavenumbers outside the sample
    range are filled with boundary values (no extrapolation).
    """
    # Ensure sample is ascending for interp1d
    if wn_sample[0] > wn_sample[-1]:
        wn_sample = wn_sample[::-1]
        ab_sample = ab_sample[::-1]
    f = interp1d(wn_sample, ab_sample, kind="linear", bounds_error=False,
                 fill_value=(ab_sample[0], ab_sample[-1]))
    return f(wn_ref)


def fit_carbonate(
    spectrum_csv_path: Union[str, Path],
    thickness_cm: float,
    fit_range: Tuple[float, float] = (1350.0, 1800.0),
    model: str = "fixed",
    fringe_correction: str = "auto",
    save_figure: bool = True,
    figure_path: Optional[Union[str, Path]] = None,
) -> Dict:


    """Fit CO₃²⁻ carbonate doublet, replicating CO2_bestfit.xls Solver.

    # Design notes — CO3 fitting
    # ---------------------------------------------------------------
    # Three models are available (set via the `model` parameter):
    #
    # model="fixed"  (default, matches Excel CO2_bestfit.xls)
    #   5-component unconstrained lstsq over fit_range:
    #   y = c0 + c1*E_bkg + c2*F_1630 + c3*G_CO3 + c4*H_line
    #   E_bkg = WOK162 devitalized glass reference spectrum.
    #   Limitation: if sample composition differs from WOK162, E_bkg
    #   may not represent the background well, giving poor R^2.
    #
    # model="taylor"
    #   Fixed baseline (E_bkg + H_line) with Taylor-expanded Gaussians
    #   for independent CO3 peak shifts.
    #   9-component bounded lsq_linear:
    #   y = c0 + c1*E_bkg + c2*F_1630 + a1430*G1430 + d1430*G1430'
    #       + a1515*G1515 + d1515*G1515' + c4*H_line + offset
    #   Recovers shift_1430 and shift_1515 independently.
    #   If |shift| > 15 cm-1: quality flag becomes "SUSPECT".
    #
    # model="pca_shift"
    #   PCA baseline + Taylor-expanded Gaussians for independent peak shifts.
    #   14-component bounded lsq_linear (adds G1430', G1515' derivatives).
    #   Recovers shift_1430 and shift_1515 independently.
    #   If |shift| > 15 cm-1: quality flag becomes "SUSPECT" (likely not CO3).
    #   See Section 3b/3c in CLAUDE.md for details.
    #
    # All models also run a shift sweep (delta = -20..+20 cm-1)
    # to estimate CO2 uncertainty from peak position.
    #
    # Notes:
    #   - The two doublet peaks do NOT necessarily shift equally
    #     (different cation environments affect each mode differently)
    #   - For more flexibility, see PyIROGlass (independent Gaussians + MCMC)
    #   - Reference: Shi et al. 2024, doi:10.30909/vol.07.02.471501
    # ---------------------------------------------------------------

    Parameters
    ----------
    spectrum_csv_path : str or Path
        Path to spectrum CSV (wavenumber vs absorbance).
    thickness_cm : float
        Sample thickness in centimeters (column C in the Excel template).
    fit_range : (float, float)
        Wavenumber range for least-squares fitting (cm⁻¹).
        Default (1350, 1800) matches the Excel template's 116-point range.
    model : str
        "fixed"     — 5-component model with fixed CO3 reference peaks (default).
                      Matches Excel CO2_bestfit.xls Solver.
        "taylor"    — 9-component model: fixed baseline (E_bkg + H_line) with
                      Taylor-expanded Gaussians for independent CO3 peak shifts.
                      Returns extra keys: "shift_1430", "shift_1515".
        "pca_shift" — 14-component PCA baseline + Taylor-expanded Gaussians
                      with independent peak shift estimation.
                      Uses bounded linear least squares (lsq_linear).
                      Returns extra keys: "shift_1430", "shift_1515".
    save_figure : bool
        If True, save diagnostic plots.
    figure_path : str or Path or None
        Explicit path for the figure.

    Returns
    -------
    dict with keys:
        "coefficients"   – dict of fitted parameters (offset, bkg, c1630, cCO3, st)
        "r_squared"      – R² goodness of fit
        "co3_peak_1515"  – CO₃²⁻ absorbance at ~1515 cm⁻¹ (by subtraction)
        "co3_peak_1430"  – CO₃²⁻ absorbance at ~1430 cm⁻¹ (by subtraction)
        "co3_absorbance" – average of the two CO₃²⁻ peak heights
        "h2o_1630_coeff" – 1630 cm⁻¹ H₂O coefficient (c_1630)
        "residual_sum_sq" – sum of squared residuals in fit range
        "wavenumber"     – full wavenumber grid
        "data"           – sample absorbance on reference grid
        "fitted"         – full fitted spectrum
        "co3_subtracted" – CO₃²⁻ band by subtraction (T column)
        "figure_path"    – path to saved figure (or None)
    """
    # Load reference spectra
    ref = _load_reference_spectra()
    wn_ref = ref["wavenumber"]   # descending order (2198 → 1203)
    E_bkg = ref["E_bkg"]
    F_1630 = ref["F_1630"]
    G_CO3 = ref["G_CO3"]
    n_pts = len(wn_ref)

    # Straight line component: H = 1, 2, 3, ..., n_pts
    H_line = np.arange(1, n_pts + 1, dtype=float)

    # Read and interpolate sample spectrum onto reference grid
    wn_sample, ab_sample = _read_spectrum_csv(spectrum_csv_path)
    data_on_grid = _interpolate_to_reference_grid(wn_sample, ab_sample, wn_ref)

    # Normalize spectrum to 100 µm thickness (matching CO2_bestfit.xls preprocessing)
    # The Excel template normalizes: data_norm = data_raw * (100 / thickness_µm)
    # so that fitted coefficients correspond to a 100-µm path length.
    thickness_um = thickness_cm * 10000.0
    data_on_grid = data_on_grid * (100.0 / thickness_um)

    # --- Fringe period estimation (for fringe_correction) ---
    # Detect periodic fringes in 2200-2800 cm-1 of the raw spectrum,
    # excluding 2280-2400 cm-1 (atmospheric CO2 gas absorption band).
    # Multi-harmonic sin/cos basis functions (fundamental + 2nd harmonic)
    # are added to the design matrix so the fitter removes fringes jointly
    # with CO3 estimation (no extrapolation needed).
    _fringe_period_est = None
    if fringe_correction in ("auto", "always"):
        try:
            from scipy.signal import savgol_filter as _sgf_fc
            from scipy.signal import find_peaks as _fp_fc
            _fc_mask = ((wn_sample >= 2200) & (wn_sample <= 2280)) | \
                       ((wn_sample >= 2400) & (wn_sample <= 2800))
            _fc_wn = wn_sample[_fc_mask]
            _fc_ab = ab_sample[_fc_mask]
            if len(_fc_wn) >= 40:
                _fc_good = np.isfinite(_fc_ab)
                _fc_wn = _fc_wn[_fc_good]
                _fc_ab = _fc_ab[_fc_good]
                _fc_dwn = (_fc_wn.max() - _fc_wn.min()) / max(len(_fc_wn) - 1, 1)
                _fc_c = np.polyfit(_fc_wn, _fc_ab, 2)
                _fc_detr = _fc_ab - np.polyval(_fc_c, _fc_wn)
                _fc_win = int(20.0 / _fc_dwn) if _fc_dwn > 0 else 31
                _fc_win = _fc_win if _fc_win % 2 == 1 else _fc_win + 1
                _fc_win = max(5, min(_fc_win, len(_fc_detr) // 2 * 2 - 1))
                _fc_smooth = _sgf_fc(_fc_detr, _fc_win, 3)
                _fc_md = max(8, int(25.0 / _fc_dwn)) if _fc_dwn > 0 else 8
                _fc_pk, _ = _fp_fc(_fc_smooth, distance=_fc_md)
                _fc_tr, _ = _fp_fc(-_fc_smooth, distance=_fc_md)
                _fc_ext = np.sort(np.concatenate([_fc_wn[_fc_pk], _fc_wn[_fc_tr]]))
                if len(_fc_ext) > 1:
                    _fc_merged = [_fc_ext[0]]
                    for _e in _fc_ext[1:]:
                        if abs(_e - _fc_merged[-1]) < 20.0:
                            _fc_merged[-1] = (_fc_merged[-1] + _e) / 2.0
                        else:
                            _fc_merged.append(_e)
                    _fc_ext = np.array(_fc_merged)
                if len(_fc_ext) >= 4:
                    _fc_hs = np.abs(np.diff(_fc_ext))
                    _fc_mhs = np.mean(_fc_hs)
                    _fc_cv = np.std(_fc_hs) / _fc_mhs if _fc_mhs > 0 else 99.0
                    _fc_amp = (_fc_smooth.max() - _fc_smooth.min()) / 2.0
                    # Require CV < 0.35 (periodic) AND amplitude > 0.003
                    # (above noise floor, typical noise ~0.001-0.002)
                    # "always" mode: bypass detection criteria and force correction
                    if fringe_correction == "always" or \
                       (_fc_cv < 0.35 and _fc_amp > 0.003):
                        _fringe_period_est = 2.0 * _fc_mhs
        except Exception:
            pass

    # Identify fitting range mask
    fit_lo, fit_hi = min(fit_range), max(fit_range)
    fit_mask = (wn_ref >= fit_lo) & (wn_ref <= fit_hi)
    n_fit = fit_mask.sum()
    if n_fit < 10:
        raise ValueError(
            f"Only {n_fit} points in fitting range [{fit_lo}, {fit_hi}]. "
            "Check that the spectrum covers this wavenumber range."
        )

    # Build design matrix for the fitting range
    E_fit = E_bkg[fit_mask]
    F_fit = F_1630[fit_mask]
    G_fit = G_CO3[fit_mask]
    H_fit = H_line[fit_mask]
    data_fit = data_on_grid[fit_mask]

    # --- Unconstrained least-squares fitting (matching Excel CO2_bestfit.xls) ---
    # The Excel Solver uses unconstrained linear minimization (solver_lin=1).
    # We use numpy lstsq, which gives the identical mathematically optimal solution.
    #
    # Rationale for NO background <= data constraint:
    #   E_bkg is a measured reference spectrum from one specific devitalized glass.
    #   For samples with different compositions, E_bkg shape may not exactly match
    #   the sample background in the fitting range (1350-1800 cm-1).  Forcing
    #   background <= data then drives c1 (E_bkg coefficient) toward zero, which:
    #     (a) causes the background extrapolation below 1350 cm-1 to be a flat line
    #         instead of the physically correct Si-O absorption curve, and
    #     (b) can cause SLSQP to return infeasible results (e.g. c_CO3 = -3.5)
    #         for poor-quality spectra.
    #
    # Post-processing interpretation:
    #   c3 <= 0  => NEGATIVE quality flag, CO3 treated as 0 in concentrations
    #   c3 >  0  => GOOD/FAIR/POOR based on R^2, SNR, n_bkg_above_data
    shift_delta = 0.0    # only set by pca_shift model
    # Initialize fringe-applied flags for all models (set True within each branch)
    _pca_fringe_applied = False
    _taylor_fringe_applied = False
    _fixed_fringe_applied = False

    if model == "pca_shift":
        # PCA baseline + Taylor-expanded Gaussians for independent peak shifts.
        # Baseline: mean + 4 PCA components from 57 devolatilized glass spectra.
        # H2Om,1635: empirical peak + 2 PCs (replaces our F_1630 Gaussian).
        # CO3: two independent Gaussians with Taylor shift terms.
        # Adds G1430' and G1515' (Gaussian derivatives
        # w.r.t. position) as extra basis functions.
        # G'(nu; mu, sigma) = G(nu) * (nu - mu) / sigma^2
        # After fitting: shift = c_deriv / c_amp  (cm-1)
        # If |shift| > 15 cm-1: flagged as SUSPECT (likely not CO3).
        from scipy.optimize import lsq_linear

        pca = _load_pca_components(wn_ref)

        # Fitting range: 1250-2200
        pca_lo = 1250.0
        pca_hi = 2200.0
        pca_mask = (wn_ref >= pca_lo) & (wn_ref <= pca_hi)
        n_pca_fit = pca_mask.sum()
        wn_fit_pca = wn_ref[pca_mask]

        # Normalized wavenumber for linear tilt
        _pca_wn_center = float(np.mean(wn_fit_pca))
        _pca_wn_scale = float(np.std(wn_fit_pca))
        _pca_wn_n_full = (wn_ref - _pca_wn_center) / _pca_wn_scale
        _pca_wn_n_fit = _pca_wn_n_full[pca_mask]

        # CO3 Gaussians and their derivatives w.r.t. position
        _co3_sigma = 30.0
        _co3_mu1430 = 1430.0
        _co3_mu1515 = 1515.0
        G1430_fit = np.exp(-((wn_fit_pca - _co3_mu1430) ** 2) / (2 * _co3_sigma**2))
        G1515_fit = np.exp(-((wn_fit_pca - _co3_mu1515) ** 2) / (2 * _co3_sigma**2))
        # Derivative: dG/dmu = G * (nu - mu) / sigma^2
        dG1430_fit = G1430_fit * (wn_fit_pca - _co3_mu1430) / (_co3_sigma**2)
        dG1515_fit = G1515_fit * (wn_fit_pca - _co3_mu1515) / (_co3_sigma**2)

        G1430_full = np.exp(-((wn_ref - _co3_mu1430) ** 2) / (2 * _co3_sigma**2))
        G1515_full = np.exp(-((wn_ref - _co3_mu1515) ** 2) / (2 * _co3_sigma**2))
        dG1430_full = G1430_full * (wn_ref - _co3_mu1430) / (_co3_sigma**2)
        dG1515_full = G1515_full * (wn_ref - _co3_mu1515) / (_co3_sigma**2)

        # Design matrix (14 parameters, or 16 with fringe correction)
        # [0-4]   PCA baseline: B_mean, BPC1..4
        # [5-7]   H2Om: mean, PC1, PC2
        # [8]     G1430 amplitude
        # [9]     G1430' (derivative — encodes shift)
        # [10]    G1515 amplitude
        # [11]    G1515' (derivative — encodes shift)
        # [12]    linear tilt
        # [13]    offset
        # [14-17] (optional) sin/cos fringe basis (2 harmonics) at estimated period
        _pca_fringe_applied = False
        _cols = [
            pca["baseline_mean"][pca_mask],   # 0
            pca["baseline_pc1"][pca_mask],    # 1
            pca["baseline_pc2"][pca_mask],    # 2
            pca["baseline_pc3"][pca_mask],    # 3
            pca["baseline_pc4"][pca_mask],    # 4
            pca["h2om_mean"][pca_mask],       # 5
            pca["h2om_pc1"][pca_mask],        # 6
            pca["h2om_pc2"][pca_mask],        # 7
            G1430_fit,                         # 8
            dG1430_fit,                        # 9
            G1515_fit,                         # 10
            dG1515_fit,                        # 11
            _pca_wn_n_fit,                     # 12
            np.ones(n_pca_fit),                # 13
        ]
        # Full-grid fringe basis (for reconstruction)
        _fringe_sin1_full = None
        _fringe_cos1_full = None
        _fringe_sin2_full = None
        _fringe_cos2_full = None
        if _fringe_period_est is not None:
            _pca_fringe_applied = True
            # First harmonic
            _fringe_sin1_fit = np.sin(2 * np.pi * wn_fit_pca / _fringe_period_est)
            _fringe_cos1_fit = np.cos(2 * np.pi * wn_fit_pca / _fringe_period_est)
            _fringe_sin1_full = np.sin(2 * np.pi * wn_ref / _fringe_period_est)
            _fringe_cos1_full = np.cos(2 * np.pi * wn_ref / _fringe_period_est)
            # Second harmonic (captures non-sinusoidal fringe shapes)
            _fringe_sin2_fit = np.sin(4 * np.pi * wn_fit_pca / _fringe_period_est)
            _fringe_cos2_fit = np.cos(4 * np.pi * wn_fit_pca / _fringe_period_est)
            _fringe_sin2_full = np.sin(4 * np.pi * wn_ref / _fringe_period_est)
            _fringe_cos2_full = np.cos(4 * np.pi * wn_ref / _fringe_period_est)
            _cols.append(_fringe_sin1_fit)   # 14
            _cols.append(_fringe_cos1_fit)   # 15
            _cols.append(_fringe_sin2_fit)   # 16
            _cols.append(_fringe_cos2_fit)   # 17

        A_pca = np.column_stack(_cols)
        data_pca_fit = data_on_grid[pca_mask]

        # Bounds: same as pca for baseline/H2Om/tilt/offset;
        # CO3 amps >= 0; derivative coefficients unconstrained
        _pca_lb = np.array([0.0, -3.0, -2.0, -0.6, -0.3,  # baseline
                            0.0, -2.0, -2.0,                # H2Om
                            0.0, -np.inf,                    # G1430 amp, deriv
                            0.0, -np.inf,                    # G1515 amp, deriv
                            -0.5, -1.0])                     # tilt, offset
        _pca_ub = np.array([4.0, 3.0, 2.0, 0.6, 0.3,       # baseline
                            3.0, 2.0, 2.0,                   # H2Om
                            np.inf, np.inf,                   # G1430 amp, deriv
                            np.inf, np.inf,                   # G1515 amp, deriv
                            0.5, 3.0])                        # tilt, offset
        if _pca_fringe_applied:
            _pca_lb = np.concatenate([_pca_lb, [-np.inf, -np.inf, -np.inf, -np.inf]])
            _pca_ub = np.concatenate([_pca_ub, [np.inf, np.inf, np.inf, np.inf]])

        sol = lsq_linear(A_pca, data_pca_fit, bounds=(_pca_lb, _pca_ub))
        c_pca = sol.x

        _pca_a = c_pca[:5]   # baseline coefficients
        _pca_b = c_pca[5:8]  # H2Om coefficients
        _amp1430 = float(c_pca[8])
        _damp1430 = float(c_pca[9])
        _amp1515 = float(c_pca[10])
        _damp1515 = float(c_pca[11])
        _pca_m = float(c_pca[12])
        _pca_b0 = float(c_pca[13])

        # Recover independent peak shifts: delta = c_deriv / c_amp
        if _amp1430 > 1e-6:
            _shift1430 = _damp1430 / _amp1430
        else:
            _shift1430 = 0.0
        if _amp1515 > 1e-6:
            _shift1515 = _damp1515 / _amp1515
        else:
            _shift1515 = 0.0
        shift_delta = (_shift1430 + _shift1515) / 2.0  # average for compat

        # Map to common variables
        c0 = _pca_b0
        c1 = 0.0
        c2 = float(_pca_b[0])
        c3p = 0.0
        c4 = 0.0

        # Full-grid reconstruction (use shifted Gaussians for fitted curve)
        _pca_baseline_full = (
            _pca_a[0] * pca["baseline_mean"]
            + _pca_a[1] * pca["baseline_pc1"]
            + _pca_a[2] * pca["baseline_pc2"]
            + _pca_a[3] * pca["baseline_pc3"]
            + _pca_a[4] * pca["baseline_pc4"]
        )
        _pca_h2om_full = (
            _pca_b[0] * pca["h2om_mean"]
            + _pca_b[1] * pca["h2om_pc1"]
            + _pca_b[2] * pca["h2om_pc2"]
        )
        _pca_linear_full = _pca_m * _pca_wn_n_full + _pca_b0
        _pca_co3_full = (_amp1430 * G1430_full + _damp1430 * dG1430_full
                         + _amp1515 * G1515_full + _damp1515 * dG1515_full)

        # Fringe component (if applied) — 2 harmonics, 4 coefficients
        _pca_fringe_full = np.zeros(n_pts)
        if _pca_fringe_applied:
            _pca_fringe_full = (
                float(c_pca[14]) * _fringe_sin1_full
                + float(c_pca[15]) * _fringe_cos1_full
                + float(c_pca[16]) * _fringe_sin2_full
                + float(c_pca[17]) * _fringe_cos2_full
            )

        fitted_full = (_pca_baseline_full + _pca_h2om_full
                       + _pca_linear_full + _pca_co3_full + _pca_fringe_full)
        fit_no_co3 = (_pca_baseline_full + _pca_h2om_full
                      + _pca_linear_full + _pca_fringe_full)
        bkg_in_fit = fit_no_co3[pca_mask]

        c3 = (_amp1430 + _amp1515) / 2.0

        # Override fit_mask for R2/residuals
        fit_mask = pca_mask
        n_fit = n_pca_fit
        data_fit = data_pca_fit
        fit_lo, fit_hi = pca_lo, pca_hi

    elif model == "taylor":
        # Taylor model: fixed baseline (E_bkg + H_line) with Taylor-expanded
        # Gaussians for independent CO3 peak shift estimation.
        # y = offset + c_bkg*E + c_1630*F + a1430*G1430 + d1430*G1430'
        #     + a1515*G1515 + d1515*G1515' + c_st*H
        # 8 parameters solved via bounded lsq_linear.
        from scipy.optimize import lsq_linear as _lsq_linear_taylor

        _co3_sigma_t = 30.0
        _co3_mu1430_t = 1430.0
        _co3_mu1515_t = 1515.0
        wn_fit_taylor = wn_ref[fit_mask]

        # CO3 Gaussians in fitting range
        G1430_fit_t = np.exp(-((wn_fit_taylor - _co3_mu1430_t) ** 2) / (2 * _co3_sigma_t**2))
        G1515_fit_t = np.exp(-((wn_fit_taylor - _co3_mu1515_t) ** 2) / (2 * _co3_sigma_t**2))
        # Taylor derivatives: dG/dmu = G * (nu - mu) / sigma^2
        dG1430_fit_t = G1430_fit_t * (wn_fit_taylor - _co3_mu1430_t) / (_co3_sigma_t**2)
        dG1515_fit_t = G1515_fit_t * (wn_fit_taylor - _co3_mu1515_t) / (_co3_sigma_t**2)

        # Full-grid versions for reconstruction
        G1430_full_t = np.exp(-((wn_ref - _co3_mu1430_t) ** 2) / (2 * _co3_sigma_t**2))
        G1515_full_t = np.exp(-((wn_ref - _co3_mu1515_t) ** 2) / (2 * _co3_sigma_t**2))
        dG1430_full_t = G1430_full_t * (wn_ref - _co3_mu1430_t) / (_co3_sigma_t**2)
        dG1515_full_t = G1515_full_t * (wn_ref - _co3_mu1515_t) / (_co3_sigma_t**2)

        # Design matrix (8 parameters, or 12 with fringe correction):
        # [0] offset, [1] E_bkg, [2] F_1630,
        # [3] G1430 amp, [4] G1430' deriv,
        # [5] G1515 amp, [6] G1515' deriv,
        # [7] H_line
        # [8-11] (optional) sin/cos fringe basis (2 harmonics)
        _taylor_fringe_applied = False
        _taylor_fringe_full = np.zeros(n_pts)
        _t_cols = [
            np.ones(n_fit),   # 0: offset
            E_fit,            # 1: E_bkg
            F_fit,            # 2: F_1630
            G1430_fit_t,      # 3: G1430 amplitude
            dG1430_fit_t,     # 4: G1430' derivative
            G1515_fit_t,      # 5: G1515 amplitude
            dG1515_fit_t,     # 6: G1515' derivative
            H_fit,            # 7: H_line (slope)
        ]
        # Fringe basis for taylor model
        _t_fringe_sin1_full = None
        _t_fringe_cos1_full = None
        _t_fringe_sin2_full = None
        _t_fringe_cos2_full = None
        if _fringe_period_est is not None:
            _taylor_fringe_applied = True
            wn_fit_t_v = wn_ref[fit_mask]
            _t_fringe_sin1_fit = np.sin(2 * np.pi * wn_fit_t_v / _fringe_period_est)
            _t_fringe_cos1_fit = np.cos(2 * np.pi * wn_fit_t_v / _fringe_period_est)
            _t_fringe_sin2_fit = np.sin(4 * np.pi * wn_fit_t_v / _fringe_period_est)
            _t_fringe_cos2_fit = np.cos(4 * np.pi * wn_fit_t_v / _fringe_period_est)
            _t_fringe_sin1_full = np.sin(2 * np.pi * wn_ref / _fringe_period_est)
            _t_fringe_cos1_full = np.cos(2 * np.pi * wn_ref / _fringe_period_est)
            _t_fringe_sin2_full = np.sin(4 * np.pi * wn_ref / _fringe_period_est)
            _t_fringe_cos2_full = np.cos(4 * np.pi * wn_ref / _fringe_period_est)
            _t_cols.extend([_t_fringe_sin1_fit, _t_fringe_cos1_fit,
                            _t_fringe_sin2_fit, _t_fringe_cos2_fit])
        A_taylor = np.column_stack(_t_cols)

        # Bounds: offset/E_bkg/F_1630/H_line unconstrained; CO3 amps >= 0;
        # derivative coefficients unconstrained
        _t_lb = np.array([-np.inf, -np.inf, -np.inf,
                          0.0, -np.inf, 0.0, -np.inf,
                          -np.inf])
        _t_ub = np.full(8, np.inf)
        if _taylor_fringe_applied:
            _t_lb = np.concatenate([_t_lb, [-np.inf, -np.inf, -np.inf, -np.inf]])
            _t_ub = np.concatenate([_t_ub, [np.inf, np.inf, np.inf, np.inf]])

        sol_t = _lsq_linear_taylor(A_taylor, data_fit, bounds=(_t_lb, _t_ub))
        c_taylor = sol_t.x

        c0 = float(c_taylor[0])         # offset
        c1 = float(c_taylor[1])         # E_bkg
        c2 = float(c_taylor[2])         # F_1630
        _amp1430 = float(c_taylor[3])
        _damp1430 = float(c_taylor[4])
        _amp1515 = float(c_taylor[5])
        _damp1515 = float(c_taylor[6])
        c4 = float(c_taylor[7])         # H_line

        # Recover independent peak shifts
        if _amp1430 > 1e-6:
            _shift1430 = _damp1430 / _amp1430
        else:
            _shift1430 = 0.0
        if _amp1515 > 1e-6:
            _shift1515 = _damp1515 / _amp1515
        else:
            _shift1515 = 0.0
        shift_delta = (_shift1430 + _shift1515) / 2.0

        # Map c3 as average CO3 amplitude (for quality flag / concentration)
        c3 = (_amp1430 + _amp1515) / 2.0
        c3p = 0.0

        # Full-grid reconstruction
        _co3_full_t = (_amp1430 * G1430_full_t + _damp1430 * dG1430_full_t
                       + _amp1515 * G1515_full_t + _damp1515 * dG1515_full_t)
        # Fringe component (if applied)
        if _taylor_fringe_applied:
            _taylor_fringe_full = (
                float(c_taylor[8]) * _t_fringe_sin1_full
                + float(c_taylor[9]) * _t_fringe_cos1_full
                + float(c_taylor[10]) * _t_fringe_sin2_full
                + float(c_taylor[11]) * _t_fringe_cos2_full
            )
        fitted_full = c0 + c1 * E_bkg + c2 * F_1630 + _co3_full_t + c4 * H_line + _taylor_fringe_full
        fit_no_co3 = c0 + c1 * E_bkg + c2 * F_1630 + c4 * H_line + _taylor_fringe_full
        bkg_in_fit = (c0 + c1 * E_fit + c2 * F_fit + c4 * H_fit
                      + _taylor_fringe_full[fit_mask])

    else:
        # Fixed model: y = c0 + c1*E + c2*F + c3*G + c4*H
        # With optional fringe correction: + sin/cos (2 harmonics)
        _fixed_fringe_applied = False
        _fixed_fringe_full = np.zeros(n_pts)
        _f_cols = [np.ones(n_fit), E_fit, F_fit, G_fit, H_fit]
        _f_fringe_sin1_full = None
        _f_fringe_cos1_full = None
        _f_fringe_sin2_full = None
        _f_fringe_cos2_full = None
        if _fringe_period_est is not None:
            _fixed_fringe_applied = True
            wn_fit_f = wn_ref[fit_mask]
            _f_fringe_sin1_fit = np.sin(2 * np.pi * wn_fit_f / _fringe_period_est)
            _f_fringe_cos1_fit = np.cos(2 * np.pi * wn_fit_f / _fringe_period_est)
            _f_fringe_sin2_fit = np.sin(4 * np.pi * wn_fit_f / _fringe_period_est)
            _f_fringe_cos2_fit = np.cos(4 * np.pi * wn_fit_f / _fringe_period_est)
            _f_fringe_sin1_full = np.sin(2 * np.pi * wn_ref / _fringe_period_est)
            _f_fringe_cos1_full = np.cos(2 * np.pi * wn_ref / _fringe_period_est)
            _f_fringe_sin2_full = np.sin(4 * np.pi * wn_ref / _fringe_period_est)
            _f_fringe_cos2_full = np.cos(4 * np.pi * wn_ref / _fringe_period_est)
            _f_cols.extend([_f_fringe_sin1_fit, _f_fringe_cos1_fit,
                            _f_fringe_sin2_fit, _f_fringe_cos2_fit])
        A_f = np.column_stack(_f_cols)
        c_f = np.linalg.lstsq(A_f, data_fit, rcond=None)[0]
        c0, c1, c2, c3, c4 = c_f[0], c_f[1], c_f[2], c_f[3], c_f[4]
        c3p = 0.0
        if _fixed_fringe_applied:
            _fixed_fringe_full = (
                float(c_f[5]) * _f_fringe_sin1_full
                + float(c_f[6]) * _f_fringe_cos1_full
                + float(c_f[7]) * _f_fringe_sin2_full
                + float(c_f[8]) * _f_fringe_cos2_full
            )

        fitted_full = c0 + c1 * E_bkg + c2 * F_1630 + c3 * G_CO3 + c4 * H_line + _fixed_fringe_full
        fit_no_co3  = c0 + c1 * E_bkg + c2 * F_1630 + c4 * H_line + _fixed_fringe_full
        bkg_in_fit  = (c0 + c1 * E_fit + c2 * F_fit + c4 * H_fit
                       + _fixed_fringe_full[fit_mask])

    fitted_in_range = fitted_full[fit_mask]

    # --- Residuals and R² ---
    residuals = fitted_in_range - data_fit
    ss_res = np.sum(residuals ** 2)
    if model == "pca_shift":
        # Standard R² for PCA models (wider range, Excel formula doesn't apply)
        y_mean_fit = float(np.mean(data_fit))
        ss_tot = np.sum((data_fit - y_mean_fit) ** 2)
        r_squared = float(1.0 - ss_res / ss_tot) if ss_tot > 0 else 0.0
    else:
        # Excel CO2_bestfit.xls definition: SSR/SST with full-spectrum mean
        y_mean_full = float(np.mean(data_on_grid))   # full-spectrum mean (259 pts)
        ss_tot = np.sum((data_fit - y_mean_full) ** 2)
        ss_reg = np.sum((fitted_in_range - y_mean_full) ** 2)
        r_squared = float(ss_reg / ss_tot) if ss_tot > 0 else 0.0

    # --- Shift sweep: estimate CO2 uncertainty from peak position ---
    # Shift G_CO3 by delta = -20..+20 cm-1 and refit (fixed model at each)
    # to see how CO2 result varies with assumed peak position.
    # np.interp requires ascending xp, so sort wn_ref first.
    _sort_idx = np.argsort(wn_ref)
    _wn_asc = wn_ref[_sort_idx]
    _G_asc = G_CO3[_sort_idx]

    sweep_deltas = np.arange(-20, 21, 2, dtype=float)
    sweep_co3_abs = []
    for sd in sweep_deltas:
        G_shifted = np.interp(wn_ref, _wn_asc + sd, _G_asc)
        G_s_fit = G_shifted[fit_mask]
        if model == "pca_shift":
            # Refit with shifted Gaussians using bounded lsq_linear
            G1430_s = np.exp(-((wn_fit_pca - _co3_mu1430 + sd)**2) / (2 * _co3_sigma**2))
            G1515_s = np.exp(-((wn_fit_pca - _co3_mu1515 + sd)**2) / (2 * _co3_sigma**2))
            A_s = A_pca.copy()
            dG1430_s = G1430_s * (wn_fit_pca - _co3_mu1430 + sd) / (_co3_sigma**2)
            dG1515_s = G1515_s * (wn_fit_pca - _co3_mu1515 + sd) / (_co3_sigma**2)
            A_s[:, 8] = G1430_s
            A_s[:, 9] = dG1430_s
            A_s[:, 10] = G1515_s
            A_s[:, 11] = dG1515_s
            _tilt_idx, _off_idx = 12, 13
            sol_s = lsq_linear(A_s, data_pca_fit, bounds=(_pca_lb, _pca_ub))
            cs = sol_s.x
            bl_s = (cs[0] * pca["baseline_mean"] + cs[1] * pca["baseline_pc1"]
                    + cs[2] * pca["baseline_pc2"] + cs[3] * pca["baseline_pc3"]
                    + cs[4] * pca["baseline_pc4"]
                    + cs[_tilt_idx] * _pca_wn_n_full + cs[_off_idx])
            h2om_s = (cs[5] * pca["h2om_mean"] + cs[6] * pca["h2om_pc1"]
                      + cs[7] * pca["h2om_pc2"])
            _fringe_s = np.zeros(n_pts)
            if _pca_fringe_applied:
                _fringe_s = (cs[14] * _fringe_sin1_full + cs[15] * _fringe_cos1_full
                             + cs[16] * _fringe_sin2_full + cs[17] * _fringe_cos2_full)
            fit_no_co3_s = bl_s + h2om_s + _fringe_s
        elif model == "taylor":
            # Refit taylor model with shifted Gaussians
            wn_fit_t = wn_ref[fit_mask]
            G1430_s = np.exp(-((wn_fit_t - _co3_mu1430_t + sd)**2) / (2 * _co3_sigma_t**2))
            G1515_s = np.exp(-((wn_fit_t - _co3_mu1515_t + sd)**2) / (2 * _co3_sigma_t**2))
            dG1430_s = G1430_s * (wn_fit_t - _co3_mu1430_t + sd) / (_co3_sigma_t**2)
            dG1515_s = G1515_s * (wn_fit_t - _co3_mu1515_t + sd) / (_co3_sigma_t**2)
            _t_sweep_cols = [
                np.ones(n_fit), E_fit, F_fit,
                G1430_s, dG1430_s, G1515_s, dG1515_s,
                H_fit,
            ]
            if _taylor_fringe_applied:
                _t_sweep_cols.extend([_t_fringe_sin1_fit, _t_fringe_cos1_fit,
                                      _t_fringe_sin2_fit, _t_fringe_cos2_fit])
            A_s_t = np.column_stack(_t_sweep_cols)
            sol_s_t = _lsq_linear_taylor(A_s_t, data_fit, bounds=(_t_lb, _t_ub))
            cs = sol_s_t.x
            _fringe_s_t = np.zeros(n_pts)
            if _taylor_fringe_applied:
                _fringe_s_t = (cs[8] * _t_fringe_sin1_full + cs[9] * _t_fringe_cos1_full
                               + cs[10] * _t_fringe_sin2_full + cs[11] * _t_fringe_cos2_full)
            fit_no_co3_s = cs[0] + cs[1] * E_bkg + cs[2] * F_1630 + cs[7] * H_line + _fringe_s_t
        else:
            _f_sweep_cols = [np.ones(n_fit), E_fit, F_fit, G_s_fit, H_fit]
            if _fixed_fringe_applied:
                _f_sweep_cols.extend([_f_fringe_sin1_fit, _f_fringe_cos1_fit,
                                      _f_fringe_sin2_fit, _f_fringe_cos2_fit])
            A_s = np.column_stack(_f_sweep_cols)
            cs = np.linalg.lstsq(A_s, data_fit, rcond=None)[0]
            _fringe_s_f = np.zeros(n_pts)
            if _fixed_fringe_applied:
                _fringe_s_f = (cs[5] * _f_fringe_sin1_full + cs[6] * _f_fringe_cos1_full
                               + cs[7] * _f_fringe_sin2_full + cs[8] * _f_fringe_cos2_full)
            fit_no_co3_s = cs[0] + cs[1] * E_bkg + cs[2] * F_1630 + cs[4] * H_line + _fringe_s_f
        co3_sub_s = data_on_grid - fit_no_co3_s
        m1515 = (wn_ref >= 1496) & (wn_ref <= 1532)
        m1430 = (wn_ref >= 1396) & (wn_ref <= 1451)
        p1515 = float(np.max(co3_sub_s[m1515])) if m1515.any() else 0.0
        p1430 = float(np.max(co3_sub_s[m1430])) if m1430.any() else 0.0
        sweep_co3_abs.append((p1515 + p1430) / 2.0)
    sweep_co3_abs = np.array(sweep_co3_abs)
    co3_sweep_min = float(np.min(sweep_co3_abs))
    co3_sweep_max = float(np.max(sweep_co3_abs))

    # --- CO₃²⁻ band by subtraction ---
    # fit_no_co3 is set within each model block above.
    co3_subtracted = data_on_grid - fit_no_co3

    # --- Find CO₃²⁻ peak heights ---
    # Peak windows match Excel (T175:T184 and T196:T210):
    #   1515 peak: ~[1496.51, 1531.23] cm-1
    #   1430 peak: ~[1396.23, 1450.23] cm-1
    # For PCA models, tighten the 1430 window to [1410, 1451] to avoid
    # Si-O upturn residuals near the low-wavenumber edge.
    _1430_lo = 1410.0 if model == "pca_shift" else 1396.0

    # Peak near 1515 cm⁻¹
    mask_1515 = (wn_ref >= 1496) & (wn_ref <= 1532)
    if mask_1515.any():
        idx_1515 = np.argmax(co3_subtracted[mask_1515])
        co3_peak_1515 = co3_subtracted[mask_1515][idx_1515]
        wn_1515 = wn_ref[mask_1515][idx_1515]
    else:
        co3_peak_1515 = 0.0
        wn_1515 = 1515.0

    # Peak near 1430 cm⁻¹
    mask_1430 = (wn_ref >= _1430_lo) & (wn_ref <= 1451)
    if mask_1430.any():
        idx_1430 = np.argmax(co3_subtracted[mask_1430])
        co3_peak_1430 = co3_subtracted[mask_1430][idx_1430]
        wn_1430 = wn_ref[mask_1430][idx_1430]
    else:
        co3_peak_1430 = 0.0
        wn_1430 = 1430.0

    co3_absorbance = (co3_peak_1515 + co3_peak_1430) / 2.0

    # --- Quality metrics ---
    # 1. Background vs data: count fit-range points where bkg > data
    # bkg_in_fit is set within each model block above.
    n_bkg_above = int(np.sum(bkg_in_fit > data_fit + 1e-6))

    # 2. Doublet symmetry: ratio of the two CO3 peak heights (ideal ~ 1.0)
    if co3_peak_1430 > 1e-6 and co3_peak_1515 > 1e-6:
        doublet_ratio = float(co3_peak_1515 / co3_peak_1430)
    else:
        doublet_ratio = 0.0

    # 3. CO3 signal-to-noise: CO3 absorbance vs residual RMS
    residual_rms = float(np.sqrt(ss_res / n_fit))
    if residual_rms > 1e-10:
        co3_snr = float(co3_absorbance / residual_rms)
    else:
        co3_snr = float("inf")

    # 4. Composite quality flag
    #    Step 1: Base flag from R2 and SNR
    #      NEGATIVE: cCO3 <= 0 (no CO3 signal detected)
    #      GOOD:     R2 >= 0.95, SNR >= 5
    #      FAIR:     R2 >= 0.80, SNR >= 2
    #      POOR:     everything else
    #    Step 2 (taylor/pca_shift only): Peak shift penalty
    #      The maximum absolute shift of the two CO3 peaks (1430, 1515) is
    #      used to cap the quality flag. Thresholds are non-linear:
    #        |shift| <= 5  cm-1:  no penalty (normal compositional variation)
    #        |shift| 5-10  cm-1:  cap at GOOD  (minor shift, still reliable)
    #        |shift| 10-15 cm-1:  cap at FAIR  (significant shift, use caution)
    #        |shift| 15-20 cm-1:  force SUSPECT (likely not standard CO3)
    #        |shift| > 20  cm-1:  force POOR   (almost certainly not CO3)
    if c3 <= 0 or co3_absorbance <= 0:
        quality_flag = "NEGATIVE"
    elif r_squared >= 0.95 and co3_snr >= 5:
        quality_flag = "GOOD"
    elif r_squared >= 0.80 and co3_snr >= 2:
        quality_flag = "FAIR"
    else:
        quality_flag = "POOR"
    # For taylor/pca_shift: report actual peak positions from subtracted spectrum
    # (more reliable than Taylor coefficient ratios which overestimate shifts).
    if model in ("taylor", "pca_shift"):
        _shift1430 = wn_1430 - 1430.0
        _shift1515 = wn_1515 - 1515.0
        shift_delta = (_shift1430 + _shift1515) / 2.0
        # Graded shift penalty: use the larger of the two peak shifts
        max_shift = max(abs(_shift1430), abs(_shift1515))
        if quality_flag not in ("NEGATIVE",):
            _rank = {"GOOD": 3, "FAIR": 2, "SUSPECT": 1, "POOR": 0}
            if max_shift > 20.0:
                # >20 cm-1: almost certainly not CO3 doublet
                quality_flag = "POOR"
            elif max_shift > 15.0:
                # 15-20 cm-1: likely not standard dissolved CO3
                quality_flag = "SUSPECT"
            elif max_shift > 10.0:
                # 10-15 cm-1: significant shift, cap at FAIR
                if _rank.get(quality_flag, 0) > _rank["FAIR"]:
                    quality_flag = "FAIR"
            # <= 10 cm-1: no downgrade (normal compositional variation)

    # Step 2b (taylor/pca_shift only): Intensity ratio check
    # For dissolved CO3 in volcanic glass, the 1515/1430 ratio is ~0.6-1.3
    # based on 7 reference spectra (basalt-andesite, SiO2=47-57%):
    #   BLRH-2-1=1.08, BLRH-10-1=0.97, qvtc-1-0=0.64, qvtc-1-2=0.75,
    #   qvtc-1-20=1.11, PC340-30=1.53(poor fit), PC340-100=1.20
    # Extreme ratios suggest the "peaks" are fringe artefacts, not CO3.
    if model in ("taylor", "pca_shift") and quality_flag not in ("NEGATIVE",):
        if doublet_ratio > 0 and (doublet_ratio < 0.6 or doublet_ratio > 1.3):
            _rank = {"GOOD": 3, "FAIR": 2, "SUSPECT": 1, "POOR": 0}
            if _rank.get(quality_flag, 0) > _rank["SUSPECT"]:
                quality_flag = "SUSPECT"

    # Step 3: Fringe contamination check (all models)
    # Detect periodic interference fringes in the absorption-free region
    # (2200-2800 cm-1) of the RAW spectrum. If fringe amplitude is comparable
    # to the CO3 absorbance, the CO3 result may be a fringe artefact.
    #
    # Algorithm:
    #   1. Extract 2200-2800 cm-1 from the raw spectrum (before normalization)
    #   2. Detrend with degree-2 polynomial
    #   3. Savitzky-Golay smooth (window=31) to suppress noise
    #   4. Find peaks and troughs; compute half-spacings between consecutive extrema
    #   5. If >= 4 extrema and CV of half-spacings < 0.35 => periodic fringes detected
    #   6. Compare fringe amplitude to CO3 absorbance (un-normalized):
    #      fringe_ratio = fringe_amplitude / (co3_absorbance / (100/thickness_um))
    #      If fringe_ratio >= 0.5: fringes are significant relative to CO3 signal
    #   7. Quality penalty: fringe_ratio >= 1.0 => cap at SUSPECT
    #                       fringe_ratio >= 0.5 => cap at FAIR
    fringe_detected = False
    fringe_amplitude = 0.0
    fringe_period = 0.0
    fringe_ratio = 0.0
    try:
        from scipy.signal import savgol_filter as _sgf
        # Use 2200-2800 cm-1 excluding atmospheric CO2 band (2280-2400)
        _fringe_mask = ((wn_sample >= 2200) & (wn_sample <= 2280)) | \
                       ((wn_sample >= 2400) & (wn_sample <= 2800))
        _wn_fr = wn_sample[_fringe_mask]
        _ab_fr = ab_sample[_fringe_mask]
        if len(_wn_fr) >= 40:
            # Detrend
            _good = np.isfinite(_ab_fr)
            _wn_fr = _wn_fr[_good]
            _ab_fr = _ab_fr[_good]
            _c = np.polyfit(_wn_fr, _ab_fr, 2)
            _detr = _ab_fr - np.polyval(_c, _wn_fr)
            # Smooth — adaptive window based on spectral resolution.
            # Target smoothing width ~20 cm-1 to suppress noise while
            # preserving fringes (half-period >= 30 cm-1 for d <= 110 um).
            # Robust spacing estimate: use total range / (n-1) instead of
            # median(diff) to handle data with rounded wavenumbers.
            _dwn = (_wn_fr.max() - _wn_fr.min()) / max(len(_wn_fr) - 1, 1)
            _target_width_cm1 = 20.0
            _win = int(_target_width_cm1 / _dwn) if _dwn > 0 else 31
            _win = _win if _win % 2 == 1 else _win + 1  # must be odd
            _win = max(5, min(_win, len(_detr) // 2 * 2 - 1))
            if _win >= 5:
                _smooth = _sgf(_detr, _win, 3)
            else:
                _smooth = _detr
            # Find extrema with minimum distance (expect half-periods >= 30 cm-1
            # for samples up to ~110 um, so require ~25 cm-1 min spacing).
            _min_dist = max(8, int(25.0 / _dwn)) if _dwn > 0 else 8
            from scipy.signal import find_peaks as _fp
            _pk_idx, _ = _fp(_smooth, distance=_min_dist)
            _tr_idx, _ = _fp(-_smooth, distance=_min_dist)
            _all_ext = np.sort(np.concatenate([_wn_fr[_pk_idx], _wn_fr[_tr_idx]]))
            # Merge extrema that are too close together (< 20 cm-1 apart).
            # When a smoothed signal has a small plateau, find_peaks may
            # report two nearby extrema of the same type; merge them by
            # keeping the midpoint.
            if len(_all_ext) > 1:
                _merged = [_all_ext[0]]
                for _ei in range(1, len(_all_ext)):
                    if abs(_all_ext[_ei] - _merged[-1]) < 20.0:
                        _merged[-1] = (_merged[-1] + _all_ext[_ei]) / 2.0
                    else:
                        _merged.append(_all_ext[_ei])
                _all_ext = np.array(_merged)
            if len(_all_ext) >= 4:
                _half_sp = np.abs(np.diff(_all_ext))
                _mean_hs = np.mean(_half_sp)
                _cv = np.std(_half_sp) / _mean_hs if _mean_hs > 0 else 99.0
                _candidate_amp = (np.max(_smooth) - np.min(_smooth)) / 2.0
                if _cv < 0.35 and _candidate_amp > 0.003:
                    fringe_detected = True
                    fringe_amplitude = _candidate_amp
                    fringe_period = 2.0 * _mean_hs
                    # CO3 absorbance in raw (un-normalized) scale
                    _co3_raw = co3_absorbance / (100.0 / thickness_um)
                    if _co3_raw > 1e-6:
                        fringe_ratio = fringe_amplitude / _co3_raw
                    else:
                        fringe_ratio = float("inf") if fringe_amplitude > 1e-6 else 0.0
    except Exception:
        pass  # fringe check is best-effort, don't break the main flow

    if fringe_detected and quality_flag not in ("NEGATIVE",):
        _rank = {"GOOD": 3, "FAIR": 2, "SUSPECT": 1, "POOR": 0}
        if fringe_ratio >= 1.0:
            # Fringe amplitude >= CO3 amplitude: very likely artefact
            if _rank.get(quality_flag, 0) > _rank["SUSPECT"]:
                quality_flag = "SUSPECT"
        elif fringe_ratio >= 0.5:
            # Fringe amplitude >= half of CO3: use with caution
            if _rank.get(quality_flag, 0) > _rank["FAIR"]:
                quality_flag = "FAIR"

    # --- Diagnostic plot ---
    fig_saved = None
    if save_figure:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt

        fig, axes = plt.subplots(3, 1, figsize=(12, 12))

        # Panel 1: Data + Fit + Components
        ax1 = axes[0]
        ax1.plot(wn_ref, data_on_grid, "k-", lw=0.8, label="Sample data")
        ax1.plot(wn_ref, fitted_full, "r-", lw=1.2, label="Fitted model")
        if model == "pca_shift":
            # Show PCA baseline, H2Om, and two CO3 Gaussians
            ax1.plot(wn_ref, _pca_baseline_full + _pca_linear_full,
                     "g--", lw=0.6, alpha=0.7, label="PCA baseline")
            ax1.plot(wn_ref, _pca_h2om_full, "b--", lw=0.6, alpha=0.7,
                     label="H$_2$Om,1635 PCA")
            ax1.plot(wn_ref, _pca_co3_full, "m--", lw=0.6, alpha=0.7,
                     label=f"CO$_3^{{2-}}$ (1430={_amp1430:.4f}, 1515={_amp1515:.4f})")
            # ax1.plot(wn_ref, fit_no_co3, "g-", lw=1.0, alpha=0.8,
            #          label="Background (excl. CO$_3^{2-}$)")
        elif model == "taylor":
            ax1.plot(wn_ref, c1 * E_bkg + c0, "g--", lw=0.6, alpha=0.7,
                     label=f"Background (c_bkg={c1:.4f})")
            ax1.plot(wn_ref, c2 * F_1630, "b--", lw=0.6, alpha=0.7,
                     label=f"1630 H$_2$O (c_1630={c2:.4f})")
            ax1.plot(wn_ref, _co3_full_t, "m--", lw=0.6, alpha=0.7,
                     label=f"CO$_3^{{2-}}$ (1430={_amp1430:.4f}, 1515={_amp1515:.4f})")
            # ax1.plot(wn_ref, fit_no_co3, "g-", lw=1.0, alpha=0.8,
            #          label="Background (excl. CO$_3^{2-}$)")
        else:
            ax1.plot(wn_ref, c1 * E_bkg + c0, "g--", lw=0.6, alpha=0.7,
                     label=f"Background (c_bkg={c1:.4f})")
            ax1.plot(wn_ref, c2 * F_1630, "b--", lw=0.6, alpha=0.7,
                     label=f"1630 H$_2$O (c_1630={c2:.4f})")
            ax1.plot(wn_ref, c3 * G_CO3, "m--", lw=0.6, alpha=0.7,
                     label=f"CO$_3^{{2-}}$ (c_CO3={c3:.4f})")
            # ax1.plot(wn_ref, fit_no_co3, "g-", lw=1.0, alpha=0.8,
            #          label="Background (excl. CO$_3^{2-}$)")
        ax1.axvspan(fit_lo, fit_hi, color="yellow", alpha=0.1, label="Fitting range")
        ax1.set_ylabel("Absorbance")
        title_extra = ""
        if model in ("taylor", "pca_shift"):
            title_extra = (f"  |  d1430={_shift1430:+.1f}  d1515={_shift1515:+.1f} cm$^{{-1}}$")
        ax1.set_title(
            f"CO$_3^{{2-}}$ Carbonate Fit [{model}] — {Path(spectrum_csv_path).stem}\n"
            f"R$^2$ = {r_squared:.6f}{title_extra}"
        )
        ax1.legend(fontsize=8, loc="upper left")
        if model == "pca_shift":
            ax1.set_xlim(fit_hi + 50, fit_lo - 50)  # show full PCA range
        else:
            ax1.set_xlim(1800, 1300)   # focus on fitting/CO3 region
        ax1.grid(True, alpha=0.3)

        # Panel 2: CO₃²⁻ band by subtraction
        # x-axis restricted to the fitting range (1350-1800 cm⁻¹) because
        # outside this range the subtracted spectrum contains Si-O glass
        # absorption artefacts (from E_bkg mismatch) that are not relevant
        # to the CO3 analysis.  CO3 peaks are read at ~1430 and ~1515 cm⁻¹.
        p2_lo, p2_hi = 1300.0, 1900.0   # display window around fitting range
        p2_mask = (wn_ref >= p2_lo) & (wn_ref <= p2_hi)
        ax2 = axes[1]
        ax2.fill_between(wn_ref[p2_mask], 0, co3_subtracted[p2_mask],
                         color="mediumpurple", alpha=0.3)
        ax2.plot(wn_ref[p2_mask], co3_subtracted[p2_mask], "m-", lw=0.8,
                 label="CO$_3^{2-}$ by subtraction")
        ax2.axhline(0, color="gray", lw=0.5, ls="--")
        ax2.axvspan(fit_lo, fit_hi, color="yellow", alpha=0.1, label="Fitting range")
        ax2.plot(wn_1515, co3_peak_1515, "rv", ms=8)
        ax2.annotate(f"{co3_peak_1515:.4f}\n@ {wn_1515:.0f}",
                     xy=(wn_1515, co3_peak_1515),
                     xytext=(wn_1515 + 60, co3_peak_1515 * 0.8),
                     fontsize=9,
                     arrowprops=dict(arrowstyle="->", color="red"),
                     bbox=dict(boxstyle="round", fc="lightyellow", ec="gray"))
        ax2.plot(wn_1430, co3_peak_1430, "rv", ms=8)
        ax2.annotate(f"{co3_peak_1430:.4f}\n@ {wn_1430:.0f}",
                     xy=(wn_1430, co3_peak_1430),
                     xytext=(wn_1430 - 100, co3_peak_1430 * 0.8),
                     fontsize=9,
                     arrowprops=dict(arrowstyle="->", color="red"),
                     bbox=dict(boxstyle="round", fc="lightyellow", ec="gray"))
        ax2.set_ylabel("Absorbance (subtracted)")
        ax2.set_title("CO$_3^{2-}$ Band by Subtraction")
        ax2.legend(fontsize=9)
        ax2.set_xlim(p2_hi, p2_lo)   # inverted x-axis, restricted to CO3 region
        ax2.grid(True, alpha=0.3)

        # Panel 3: Residuals in fitting range
        ax3 = axes[2]
        wn_fit_region = wn_ref[fit_mask]
        ax3.plot(wn_fit_region, residuals, "k-", lw=0.5)
        ax3.fill_between(wn_fit_region, 0, residuals, color="salmon", alpha=0.3)
        ax3.axhline(0, color="gray", lw=0.5, ls="--")
        ax3.set_xlabel("Wavenumber (cm$^{-1}$)")
        ax3.set_ylabel("Residual (fit - data)")
        ax3.set_title("Fit $-$ Data")
        ax3.invert_xaxis()
        ax3.grid(True, alpha=0.3)

        plt.tight_layout()

        if figure_path is None:
            figure_path = Path(spectrum_csv_path).with_name(
                Path(spectrum_csv_path).stem + "_co3_fit.png"
            )
        fig_saved = str(figure_path)
        fig.savefig(fig_saved, dpi=150)
        plt.close(fig)

    coefficients = {
        "offset": float(c0),
        "bkg": float(c1),
        "c1630": float(c2),
        "cCO3": float(c3),
        "st": float(c4),
    }
    if model in ("taylor", "pca_shift"):
        coefficients["amp_1430"] = float(_amp1430)
        coefficients["amp_1515"] = float(_amp1515)
        coefficients["shift_1430"] = float(_shift1430)
        coefficients["shift_1515"] = float(_shift1515)
    if model == "pca_shift":
        coefficients["pca_baseline"] = [float(x) for x in _pca_a]
        coefficients["pca_h2om"] = [float(x) for x in _pca_b]
        coefficients["pca_tilt"] = float(_pca_m)
        coefficients["pca_offset"] = float(_pca_b0)

    # --- CO3 uncertainty estimation ---
    # (a) Fitting noise: residual RMS as noise estimate for peak heights.
    #     Each peak height is read from one point of the subtracted spectrum,
    #     so its uncertainty is approximately the residual noise level.
    #     The average of two peaks has unc ~ residual_rms / sqrt(2),
    #     but peaks share the same baseline so conservatively use residual_rms.
    co3_absorbance_unc = residual_rms

    # (b) Peak shift sensitivity: half-range of CO3 absorbance across
    #     the +-20 cm-1 shift sweep.
    co3_shift_unc = (co3_sweep_max - co3_sweep_min) / 2.0

    # (c) Combined uncertainty (fitting noise + shift sensitivity in quadrature)
    co3_total_unc = float(np.sqrt(co3_absorbance_unc**2 + co3_shift_unc**2))

    result_dict = {
        "coefficients": coefficients,
        "model": model,
        "r_squared": float(r_squared),
        "co3_peak_1515": float(co3_peak_1515),
        "co3_peak_1430": float(co3_peak_1430),
        "co3_absorbance": float(co3_absorbance),
        "co3_absorbance_unc": float(co3_absorbance_unc),
        "co3_shift_unc": float(co3_shift_unc),
        "co3_total_unc": float(co3_total_unc),
        "h2o_1630_coeff": float(c2),
        "residual_sum_sq": float(ss_res),
        "shift_delta": float(shift_delta),
        "co3_sweep_range": (co3_sweep_min, co3_sweep_max),
        # Quality metrics
        "n_bkg_above_data": n_bkg_above,
        "doublet_ratio": doublet_ratio,
        "co3_snr": co3_snr,
        "residual_rms": residual_rms,
        "quality_flag": quality_flag,
        # Arrays
        "wavenumber": wn_ref,
        "data": data_on_grid,
        "fitted": fitted_full,
        "co3_subtracted": co3_subtracted,
        "figure_path": fig_saved,
        # taylor/pca_shift specific
        "shift_1430": float(_shift1430) if model in ("taylor", "pca_shift") else None,
        "shift_1515": float(_shift1515) if model in ("taylor", "pca_shift") else None,
        # Fringe contamination check
        "fringe_detected": fringe_detected,
        "fringe_amplitude": float(fringe_amplitude),
        "fringe_period": float(fringe_period),
        "fringe_ratio": float(fringe_ratio),
        # Fringe correction (all models when fringe_correction != "never")
        "fringe_corrected": bool(
            (_pca_fringe_applied if model == "pca_shift" else False)
            or (_taylor_fringe_applied if model == "taylor" else False)
            or (_fixed_fringe_applied if model == "fixed" else False)
        ),
        "fringe_period_est": float(_fringe_period_est) if _fringe_period_est else 0.0,
    }
    return result_dict


# ---------------------------------------------------------------------------
# Synthetic test spectrum generator (for validation when no real CSVs exist)
# ---------------------------------------------------------------------------

def _generate_synthetic_spectrum(
    output_path: Union[str, Path],
    peak_center: float = 3550.0,
    peak_amplitude: float = 0.45,
    peak_width: float = 120.0,
    baseline_slope: float = 2e-5,
    baseline_offset: float = 0.08,
    noise_level: float = 0.002,
    wn_range: Tuple[float, float] = (2000, 5000),
    n_points: int = 1500,
    seed: int = 42,
) -> Path:
    """Create a synthetic FTIR spectrum CSV for testing.

    The spectrum is a Gaussian peak on a sloped baseline with noise.
    """
    rng = np.random.default_rng(seed)
    wn = np.linspace(wn_range[0], wn_range[1], n_points)

    # Sloped baseline
    baseline = baseline_offset + baseline_slope * (wn - wn_range[0])

    # Gaussian O-H peak
    peak = peak_amplitude * np.exp(-0.5 * ((wn - peak_center) / peak_width) ** 2)

    # Add a smaller secondary peak (like a shoulder near 3200 cm⁻¹)
    shoulder = 0.08 * np.exp(-0.5 * ((wn - 3200.0) / 80.0) ** 2)

    # Noise
    noise = rng.normal(0, noise_level, n_points)

    absorbance = baseline + peak + shoulder + noise

    output_path = Path(output_path)
    with open(output_path, "w") as f:
        f.write("wavenumber,absorbance\n")
        for w, a in zip(wn, absorbance):
            f.write(f"{w:.2f},{a:.6f}\n")

    return output_path


def _generate_synthetic_co3_spectrum(
    output_path: Union[str, Path],
    co3_scale: float = 0.15,
    bkg_scale: float = 0.40,
    h2o_1630_scale: float = 0.004,
    offset: float = 0.23,
    slope: float = -0.0005,
    noise_level: float = 0.001,
    seed: int = 42,
) -> Path:
    """Create a synthetic CO₃²⁻ spectrum for testing fit_carbonate.

    Builds a spectrum using the actual reference spectra from
    co2_reference_spectra.json with known coefficients.
    """
    ref = _load_reference_spectra()
    wn = ref["wavenumber"]
    E = ref["E_bkg"]
    F = ref["F_1630"]
    G = ref["G_CO3"]
    H = np.arange(1, len(wn) + 1, dtype=float)

    rng = np.random.default_rng(seed)
    noise = rng.normal(0, noise_level, len(wn))

    ab = offset + bkg_scale * E + h2o_1630_scale * F + co3_scale * G + slope * H + noise

    output_path = Path(output_path)
    with open(output_path, "w") as f:
        f.write("wavenumber,absorbance\n")
        for w, a in zip(wn, ab):
            f.write(f"{w:.2f},{a:.6f}\n")

    return output_path


# ---------------------------------------------------------------------------
# Module 5 — Concentration Calculations (replaces Master Table formulas)
# ---------------------------------------------------------------------------
# Exact formulas from Master Table Template.xlsx row 5:
#
#   Q (Tot H2O wt%):
#     =1000000*((D5*18.02/(62.3*O5*K5)))
#     → wt% = 1e6 * A_OH * MW_H2O / (eps_OH * rho_kgm3 * d_um)
#
#   R (H2Omol 1630 wt%):
#     =1000000*((E5*18.02/(25*O5*100)))
#     → wt% = 1e6 * c_1630 * MW_H2O / (eps_1630 * rho_kgm3 * 100)
#     Note: c_1630 is from fit_carbonate (spectrum normalized to 100 µm),
#     so the denominator uses 100 (µm) instead of actual thickness.
#
#   V (CO2 ppm):
#     =10000000000*44*F5/(375*O5*100)
#     → ppm = 1e10 * MW_CO2 * c_CO3 / (eps_CO3 * rho_kgm3 * 100)
#     Same normalization convention as R.
#
# Beer-Lambert derivation:
#   A = eps * c * d_cm  →  c = A / (eps * d_cm)  [mol/L]
#   wt% = c * MW / rho_g/L * 100 = A * MW * 100 / (eps * d_cm * rho_kg/m3)
#   Since d_cm = d_um / 10000:
#   wt% = A * MW * 1e6 / (eps * d_um * rho_kg/m3)
#   ppm = wt% * 1e4
# ---------------------------------------------------------------------------

# Default molar absorptivity coefficients (L/(mol*cm))
_DEFAULT_ABSORPTIVITIES = {
    "oh_3570": 62.3,       # O-H stretch at 3570 cm-1
    "h2omol_1630": 25.0,   # molecular H2O bending at 1630 cm-1
    "co3": 375.0,          # CO3^2- doublet (1515/1430 cm-1)
}

# ---------------------------------------------------------------------------
# Composition-dependent molar absorptivity (Shi et al. 2024, PyIRoGlass)
# ---------------------------------------------------------------------------
# epsilon = m0 + m1 * compositional_parameter
#
# H2O species: parameter = tau = (Si4+ + Al3+) / total_cations
#   - Cation count per formula unit:
#       SiO2 -> 1 Si, Al2O3 -> 2 Al, Fe2O3 -> 2 Fe3+, FeO -> 1 Fe2+,
#       MgO -> 1 Mg, CaO -> 1 Ca, Na2O -> 2 Na, K2O -> 2 K,
#       TiO2 -> 1 Ti, MnO -> 1 Mn, P2O5 -> 2 P, H2O -> 2 H
#   - tau = (n_Si + n_Al) / n_total  [computed from mole fractions]
#
# CO2 species: parameter = eta = Na+ / (Na+ + Ca2+)
#   - eta = 2*X_Na2O / (2*X_Na2O + X_CaO)  [from mole fractions]
#
# Regression coefficients from PyIRoGlass core.py lines 1410-1419:
_EPSILON_REGRESSION = {
    # (m0, m1, covm_diag) -- covm_diag = diagonal of posterior covariance
    "H2Ot_3550": (15.73722543, 71.39668681, (38.05316054, 77.3885357)),
    "H2Om_1635": (-50.3975642, 124.2505339, (20.85034888, 39.38749563)),
    "CO3":       (417.17390625, -318.09377591, (84.84230954, 339.64346778)),
}

# Calibration ranges from PyIRoGlass core.py lines 1428-1434:
_TAU_RANGE_3550 = (0.508, 0.902)
_TAU_RANGE_1635 = (0.627, 0.860)
_ETA_RANGE_CO3  = (0.232, 0.841)


def calculate_epsilon(
    composition: Dict[str, float],
    fe3_over_fetotal: float = 0.15,
    mole_fractions: Optional[Dict[str, float]] = None,
) -> Dict[str, float]:
    """Calculate composition-dependent molar absorptivities.

    Uses the Shi et al. (2024) PyIRoGlass regression:
      epsilon_H2Ot_3550 = 15.737 + 71.397 * tau
      epsilon_H2Om_1635 = -50.398 + 124.251 * tau
      epsilon_CO3       = 417.174 - 318.094 * eta

    where tau = (Si4+ + Al3+) / total_cations  (from oxide mole fractions)
    and   eta = Na+ / (Na+ + Ca2+)

    Parameters
    ----------
    composition : dict
        Oxide wt% values (same format as calculate_density).
    fe3_over_fetotal : float
        Fe3+/Fe_total ratio (default 0.15).
    mole_fractions : dict or None
        Pre-computed mole fractions (from calculate_density output).
        If None, computed internally from composition.

    Returns
    -------
    dict with keys:
        "tau"                   -- tetrahedral cation fraction
        "eta"                   -- Na/(Na+Ca) cation fraction
        "epsilon_H2Ot_3550"     -- molar absorptivity for total H2O
        "sigma_epsilon_H2Ot_3550" -- 68% uncertainty
        "epsilon_H2Om_1635"     -- molar absorptivity for molecular H2O 1635
        "sigma_epsilon_H2Om_1635" -- 68% uncertainty
        "epsilon_CO3"           -- molar absorptivity for CO3 doublet
        "sigma_epsilon_CO3"     -- 68% uncertainty
        "warnings"              -- list of warning strings (empty if OK)
    """
    warnings_list = []

    # Get mole fractions
    if mole_fractions is None:
        comp = _split_iron(composition, fe3_over_fetotal)
        moles = {}
        for oxide, wt in comp.items():
            if oxide in _MOLECULAR_WEIGHTS and wt > 0:
                moles[oxide] = wt / _MOLECULAR_WEIGHTS[oxide]
        total_moles = sum(moles.values())
        if total_moles == 0:
            raise ValueError("Total moles is zero")
        mf = {ox: m / total_moles for ox, m in moles.items()}
    else:
        mf = mole_fractions

    # Cation counts from mole fractions:
    # Each oxide contributes cations: SiO2->1, TiO2->1, Al2O3->2, Fe2O3->2,
    # FeO->1, MnO->1, MgO->1, CaO->1, Na2O->2, K2O->2, P2O5->2, H2O->2
    _cation_mult = {
        "SiO2": 1, "TiO2": 1, "Al2O3": 2, "Fe2O3": 2,
        "FeO": 1, "MnO": 1, "MgO": 1, "CaO": 1,
        "Na2O": 2, "K2O": 2, "P2O5": 2, "H2O": 2,
    }
    n_cations = {}
    for oxide, mult in _cation_mult.items():
        n_cations[oxide] = mf.get(oxide, 0.0) * mult

    total_cations = sum(n_cations.values())
    if total_cations <= 0:
        raise ValueError("Total cations is zero")

    # tau = (Si4+ + Al3+) / total_cations
    n_Si = n_cations.get("SiO2", 0.0)
    n_Al = n_cations.get("Al2O3", 0.0)
    tau = (n_Si + n_Al) / total_cations

    # eta = Na+ / (Na+ + Ca2+)
    n_Na = n_cations.get("Na2O", 0.0)
    n_Ca = n_cations.get("CaO", 0.0)
    if (n_Na + n_Ca) > 0:
        eta = n_Na / (n_Na + n_Ca)
    else:
        eta = 0.5  # fallback
        warnings_list.append("Na2O and CaO both zero; eta set to 0.5")

    # Calculate epsilon values
    m0_3550, m1_3550, cov_3550 = _EPSILON_REGRESSION["H2Ot_3550"]
    m0_1635, m1_1635, cov_1635 = _EPSILON_REGRESSION["H2Om_1635"]
    m0_co3, m1_co3, cov_co3 = _EPSILON_REGRESSION["CO3"]

    eps_3550 = m0_3550 + m1_3550 * tau
    eps_1635 = m0_1635 + m1_1635 * tau
    eps_co3 = m0_co3 + m1_co3 * eta

    # Uncertainty: Equation 8 from Shi et al. (2024)
    # c_T = Z * c_m * Z^T + m * c_z * m^T
    # where c_z accounts for 1% analytical uncertainty on tau/eta
    def _sigma_epsilon(m0_cov, m1_cov, m_vec, z_val):
        # c_m contribution: [1, z] * diag(cov) * [1, z]^T
        cm_term = m0_cov * 1.0 + m1_cov * z_val**2
        # c_z contribution: [m0, m1] * diag(0, (z*0.01)^2) * [m0, m1]^T
        cz_term = m_vec[1]**2 * (z_val * 0.01)**2
        return float(np.sqrt(cm_term + cz_term))

    sigma_3550 = _sigma_epsilon(cov_3550[0], cov_3550[1],
                                (m0_3550, m1_3550), tau)
    sigma_1635 = _sigma_epsilon(cov_1635[0], cov_1635[1],
                                (m0_1635, m1_1635), tau)
    sigma_co3 = _sigma_epsilon(cov_co3[0], cov_co3[1],
                               (m0_co3, m1_co3), eta)

    # Range warnings
    if not (_TAU_RANGE_3550[0] <= tau <= _TAU_RANGE_3550[1]):
        warnings_list.append(
            f"tau={tau:.4f} outside calibration range "
            f"{_TAU_RANGE_3550} for epsilon_H2Ot_3550")
    if not (_TAU_RANGE_1635[0] <= tau <= _TAU_RANGE_1635[1]):
        warnings_list.append(
            f"tau={tau:.4f} outside calibration range "
            f"{_TAU_RANGE_1635} for epsilon_H2Om_1635")
    if not (_ETA_RANGE_CO3[0] <= eta <= _ETA_RANGE_CO3[1]):
        warnings_list.append(
            f"eta={eta:.4f} outside calibration range "
            f"{_ETA_RANGE_CO3} for epsilon_CO3")

    return {
        "tau": tau,
        "eta": eta,
        "epsilon_H2Ot_3550": eps_3550,
        "sigma_epsilon_H2Ot_3550": sigma_3550,
        "epsilon_H2Om_1635": eps_1635,
        "sigma_epsilon_H2Om_1635": sigma_1635,
        "epsilon_CO3": eps_co3,
        "sigma_epsilon_CO3": sigma_co3,
        "warnings": warnings_list,
    }


def calculate_concentration(
    oh_absorbance: float,
    h2omol_1630_coeff: float,
    co3_coeff: float,
    density_kgm3: float,
    thickness_um: float,
    epsilon_oh: float = 62.3,
    epsilon_h2omol: float = 25.0,
    epsilon_co3: float = 375.0,
    mw_h2o: float = 18.02,
    mw_co2: float = 44.0,
    oh_absorbance_unc: float = 0.0,
    co3_coeff_unc: float = 0.0,
    thickness_unc_um: float = 0.0,
    epsilon_oh_unc: float = 0.0,
    epsilon_h2omol_unc: float = 0.0,
    epsilon_co3_unc: float = 0.0,
    co3_residual_rms: float = 0.0,
    co3_snr: float = float("inf"),
) -> Dict[str, float]:
    """Calculate H2O and CO2 concentrations using Beer-Lambert law.

    Parameters
    ----------
    oh_absorbance : float
        Baseline-corrected OH peak absorbance at ~3570 cm-1
        (from fit_h2o_peak, column D in Master Table).
    h2omol_1630_coeff : float
        c_1630 coefficient from fit_carbonate (normalized to 100 um).
        This is the H2Omol 1630 absorbance as it would appear in a
        100-um-thick sample (column E in Master Table).
    co3_coeff : float
        c_CO3 coefficient from fit_carbonate (normalized to 100 um).
        This is the CO3 absorbance for a 100-um sample
        (column F in Master Table).
    density_kgm3 : float
        Glass density in kg/m3 (column O in Master Table).
    thickness_um : float
        Sample thickness in micrometers (column K in Master Table).
        Used only for the total H2O calculation from the 3570 peak.
    epsilon_oh : float
        Molar absorptivity for OH at 3570 cm-1 (default 62.3).
    epsilon_h2omol : float
        Molar absorptivity for molecular H2O at 1630 cm-1 (default 25).
    epsilon_co3 : float
        Molar absorptivity for CO3^2- (default 375).
    mw_h2o : float
        Molecular weight of H2O (default 18.02).
    mw_co2 : float
        Molecular weight of CO2 (default 44.0).
    oh_absorbance_unc : float
        Uncertainty (1-sigma) on OH absorbance from baseline variation
        (from fit_h2o_peak "peak_height_std").  Default 0 (no error).
    co3_coeff_unc : float
        Uncertainty (1-sigma) on CO3 absorbance from fitting + shift
        (from fit_carbonate "co3_total_unc").  Default 0 (no error).
    thickness_unc_um : float
        Uncertainty (1-sigma) on thickness in um from multiple fringe
        measurements (from calculate_thickness_multi "stdev_um").
        Default 0 (no error).
    epsilon_oh_unc : float
        Uncertainty (1-sigma) on epsilon_oh (default 0).
    epsilon_h2omol_unc : float
        Uncertainty (1-sigma) on epsilon_h2omol (default 0).
    epsilon_co3_unc : float
        Uncertainty (1-sigma) on epsilon_co3 (default 0).

    Returns
    -------
    dict with keys:
        "total_h2o_wt_pct"      – Total H2O wt% from OH 3570 (Master Table Q)
        "h2omol_1630_wt_pct"    – Molecular H2O wt% from 1630 (Master Table R)
        "co2_ppm"               – CO2 in ppm from CO3 doublet (Master Table V)
        "total_h2o_unc_wt_pct"  – 1-sigma uncertainty on total H2O (wt%)
        "co2_unc_ppm"           – 1-sigma uncertainty on CO2 (ppm)
    """
    if density_kgm3 <= 0:
        raise ValueError("density_kgm3 must be positive")
    if thickness_um <= 0:
        raise ValueError("thickness_um must be positive")

    # Q: Total H2O wt% = 1e6 * A_OH * MW_H2O / (eps_OH * rho * d_um)
    total_h2o = 1e6 * oh_absorbance * mw_h2o / (
        epsilon_oh * density_kgm3 * thickness_um
    )

    # R: H2Omol 1630 wt% = 1e6 * c_1630 * MW_H2O / (eps_1630 * rho * 100)
    # The 100 is the normalization thickness in um from fit_carbonate
    h2omol_1630 = 1e6 * h2omol_1630_coeff * mw_h2o / (
        epsilon_h2omol * density_kgm3 * 100.0
    )

    # V: CO2 ppm = 1e10 * MW_CO2 * c_CO3 / (eps_CO3 * rho * 100)
    co2_ppm = 1e10 * mw_co2 * co3_coeff / (
        epsilon_co3 * density_kgm3 * 100.0
    )

    # --- Uncertainty propagation ---
    # For y = k * A / (eps * d):
    #   sigma_y/y = sqrt((sigma_A/A)^2 + (sigma_d/d)^2 + (sigma_eps/eps)^2)
    rel_A_oh = (oh_absorbance_unc / oh_absorbance) if oh_absorbance > 1e-10 else 0.0
    rel_d = (thickness_unc_um / thickness_um) if thickness_um > 1e-10 else 0.0
    rel_eps_oh = (epsilon_oh_unc / epsilon_oh) if epsilon_oh > 1e-10 else 0.0
    total_h2o_unc = abs(total_h2o) * np.sqrt(
        rel_A_oh**2 + rel_d**2 + rel_eps_oh**2)

    # CO2 uncertainty: from CO3 absorbance + epsilon
    rel_co3 = (co3_coeff_unc / co3_coeff) if co3_coeff > 1e-10 else 0.0
    rel_eps_co3 = (epsilon_co3_unc / epsilon_co3) if epsilon_co3 > 1e-10 else 0.0
    co2_unc = abs(co2_ppm) * np.sqrt(rel_co3**2 + rel_eps_co3**2)

    # --- CO2 detection limit (3-sigma criterion) ---
    # Minimum detectable CO3 absorbance = 3 * residual_rms
    # Convert to CO2 ppm via Beer-Lambert (same formula as co2_ppm)
    co2_below_detection = False
    co2_detection_limit_ppm = 0.0
    if co3_residual_rms > 0:
        co3_detection_limit = 3.0 * co3_residual_rms
        co2_detection_limit_ppm = float(
            1e10 * mw_co2 * co3_detection_limit / (epsilon_co3 * density_kgm3 * 100.0)
        )
        co2_below_detection = co3_snr < 3.0

    return {
        "total_h2o_wt_pct": total_h2o,
        "h2omol_1630_wt_pct": h2omol_1630,
        "co2_ppm": co2_ppm,
        "total_h2o_unc_wt_pct": float(total_h2o_unc),
        "co2_unc_ppm": float(co2_unc),
        "co2_detection_limit_ppm": co2_detection_limit_ppm,
        "co2_below_detection": co2_below_detection,
    }


# ---------------------------------------------------------------------------
# Main Pipeline — process_sample
# ---------------------------------------------------------------------------

def process_sample(
    spectrum_csv: Union[str, Path],
    composition_dict: Dict[str, float],
    thickness_measurements: Optional[List[Tuple[float, float, int]]] = None,
    olivine_fo_a: Optional[float] = None,
    thickness_um: Optional[float] = None,
    thickness_unc_um: Optional[float] = None,
    reflectance_csv: Optional[Union[str, Path]] = None,
    olivine_fo: Optional[float] = None,
    fe3_ratio: float = 0.15,
    h2o_baseline_range: Tuple[float, float] = (2200, 3800),
    h2o_baseline_low: Tuple[float, float] = (2200, 2400),
    h2o_baseline_high: Optional[Tuple[float, float]] = None,
    co3_fit_range: Tuple[float, float] = (1350.0, 1800.0),
    co3_model: str = "fixed",
    fringe_correction: str = "auto",
    epsilon_oh: float = 62.3,
    epsilon_h2omol: float = 25.0,
    epsilon_co3: float = 375.0,
    use_composition_epsilon: bool = False,
    save_figures: bool = True,
    output_dir: Optional[Union[str, Path]] = None,
) -> Dict:
    """Full FTIR processing pipeline for a single sample.

    Chains all modules:
      1. Thickness from reflectance interference fringes
      2. Glass density and refractive index from oxide composition
      3. H2O peak baseline correction (~3570 cm-1)
      4. CO3 carbonate doublet fitting (~1430/1515 cm-1)
      5. Concentration calculations (H2O wt%, CO2 ppm)

    Three thickness input modes (mutually exclusive, checked in order):

    Mode A – Manual fringe measurements:
        Provide ``thickness_measurements`` (list of tuples).

    Mode B – Direct thickness value:
        Provide ``thickness_um`` (and optionally ``thickness_unc_um``).

    Mode C – Auto-detect from reflectance spectrum:
        Provide ``reflectance_csv`` and ``olivine_fo``.
        The refractive index is computed from the olivine Fo content and
        thickness is automatically measured from interference fringes.

    Parameters
    ----------
    spectrum_csv : str or Path
        Path to the FTIR spectrum CSV file (wavenumber vs absorbance).
    composition_dict : dict
        Glass oxide composition in wt% (e.g. {"SiO2": 50.3, "TiO2": 2.5, ...}).
        Accepts "FeOT" for total iron or separate "Fe2O3"/"FeO".
    thickness_measurements : list of (wavenumber_high, wavenumber_low, num_fringes)
        Mode A: Reflectance interference fringe measurements for thickness calc.
    thickness_um : float or None
        Mode B: Directly specify thickness in micrometers.
    thickness_unc_um : float or None
        Mode B: Uncertainty in thickness (um). Default 0 if not provided.
    reflectance_csv : str or Path or None
        Mode C: Path to reflectance spectrum file.
    olivine_fo : float or None
        Mode C: Host olivine forsterite content (0-1 or 0-100 mol%).
    fe3_ratio : float
        Fe3+/Fe_total ratio for iron splitting (default 0.15).
    h2o_baseline_range : (float, float)
        Data extraction range for H2O baseline correction (default 2200-3800).
        Must cover both the ~3550 cm-1 OH peak and the baseline anchor regions.
    h2o_baseline_low : (float, float)
        Low-wavenumber baseline anchor range (default 2200-2400 cm-1).
    h2o_baseline_high : (float, float) or None
        High-wavenumber baseline anchor range.
        Default None -> (3700, 3800) cm-1.
    co3_fit_range : (float, float)
        Wavenumber range for CO3 fitting (default 1350-1800).
        Ignored when co3_model="pca_shift" (uses 1250-2200 internally).
    co3_model : str
        CO3 fitting model: "fixed" (default), "taylor", or "pca_shift".
    epsilon_oh : float
        Molar absorptivity for OH at 3570 cm-1 (default 62.3).
    epsilon_h2omol : float
        Molar absorptivity for molecular H2O at 1630 cm-1 (default 25).
    epsilon_co3 : float
        Molar absorptivity for CO3^2- (default 375).
        Ignored if use_composition_epsilon=True.
    use_composition_epsilon : bool
        If True, compute composition-dependent epsilon values from
        Shi et al. (2024) PyIRoGlass regressions (tau/eta parameterization)
        and propagate their uncertainties. Overrides epsilon_oh/h2omol/co3.
        Default False (uses fixed epsilon values).
    save_figures : bool
        Whether to save diagnostic plots (default True).
    output_dir : str or Path or None
        Directory for output figures. If None, uses the spectrum file's directory.

    Returns
    -------
    dict with keys:
        "sample_name"       - stem of the spectrum filename
        "thickness"         - dict from thickness calculation
        "density"           - dict from calculate_density
        "refractive_index"  - dict from get_refractive_index
        "h2o_peak"          - dict from fit_h2o_peak
        "co3_fit"           - dict from fit_carbonate
        "concentration"     - dict from calculate_concentration
        "epsilon"           - dict from calculate_epsilon (if use_composition_epsilon)
    """
    spectrum_csv = Path(spectrum_csv)
    sample_name = spectrum_csv.stem

    if output_dir is None:
        output_dir = spectrum_csv.parent
    else:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

    # --- Module 2: Density & glass refractive index ---
    density_result = calculate_density(composition_dict, fe3_over_fetotal=fe3_ratio)
    ri_result = get_refractive_index(composition_dict, fe3_over_fetotal=fe3_ratio)

    # --- Module 1: Thickness (three modes) ---
    if thickness_measurements is not None:
        # Mode A: manual fringe measurements (olivine Fo required)
        if olivine_fo_a is None:
            raise ValueError("Mode A requires olivine_fo_a to calculate refractive index.")
        refractive_index = olivine_refractive_index(olivine_fo_a)
        thickness_result = calculate_thickness_multi(
            thickness_measurements, refractive_index=refractive_index
        )
    elif thickness_um is not None:
        # Mode B: direct thickness value
        _unc = thickness_unc_um if thickness_unc_um is not None else 0.0
        thickness_result = {
            "average_um": float(thickness_um),
            "stdev_um": float(_unc),
            "average_cm": float(thickness_um) / 10000.0,
            "method": "manual",
        }
    elif reflectance_csv is not None and olivine_fo is not None:
        # Mode C: auto-detect from reflectance spectrum + olivine Fo
        n_olivine = olivine_refractive_index(olivine_fo)
        fringe_fig = (output_dir / f"{sample_name}_fringes.png"
                      if save_figures else None)
        thickness_result = calculate_thickness_from_spectrum(
            reflectance_csv, refractive_index=n_olivine,
            save_figure=save_figures, figure_path=fringe_fig,
        )
        thickness_result["method"] = "auto_reflectance"
        thickness_result["olivine_fo"] = (olivine_fo / 100.0
                                          if olivine_fo > 1.0 else olivine_fo)
    else:
        raise ValueError(
            "Must provide one of: (1) thickness_measurements, "
            "(2) thickness_um, or (3) reflectance_csv + olivine_fo.")

    _thickness_um = thickness_result["average_um"]
    _thickness_cm = thickness_result["average_cm"]

    # --- Module 3: H2O baseline correction & peak height at ~3550 cm-1 ---
    h2o_fig = output_dir / f"{sample_name}_h2o_baseline.png" if save_figures else None
    h2o_kwargs = dict(
        peak_range=h2o_baseline_range,
        baseline_low_range=h2o_baseline_low,
        compute_uncertainty=True,
        save_figure=save_figures,
        figure_path=h2o_fig,
    )
    if h2o_baseline_high is not None:
        h2o_kwargs["baseline_high_range"] = h2o_baseline_high
    h2o_result = fit_h2o_peak(spectrum_csv, **h2o_kwargs)

    # --- Module 4: CO3 carbonate fitting ---
    co3_fig = output_dir / f"{sample_name}_co3_fit.png" if save_figures else None
    co3_result = fit_carbonate(
        spectrum_csv,
        thickness_cm=_thickness_cm,
        fit_range=co3_fit_range,
        model=co3_model,
        fringe_correction=fringe_correction,
        save_figure=save_figures,
        figure_path=co3_fig,
    )

    # --- Module 5: Concentrations (with density-H2O iteration) ---
    #
    # H2O wt% depends on density, but density depends on H2O content.
    # We iterate: start with H2O=0, compute density -> H2O -> update
    # composition with H2O -> recompute density -> recompute H2O, until
    # convergence (typically 2-3 iterations).
    #
    # CO3/CO2 is also recomputed each iteration since it uses density,
    # but its feedback effect on density is negligible.

    MAX_ITER = 10
    CONV_THRESHOLD = 0.001  # wt% H2O convergence threshold

    h2o_estimate = 0.0  # initial guess: no H2O
    epsilon_result = None

    for _iter in range(MAX_ITER):
        # Update composition with current H2O estimate
        comp_with_h2o = dict(composition_dict)
        if h2o_estimate > 0:
            comp_with_h2o["H2O"] = h2o_estimate

        # Recompute density with H2O in composition
        density_result = calculate_density(
            comp_with_h2o, fe3_over_fetotal=fe3_ratio)

        # Epsilon (may depend on composition/mole fractions)
        _eps_oh = epsilon_oh
        _eps_h2omol = epsilon_h2omol
        _eps_co3 = epsilon_co3
        _eps_oh_unc = 0.0
        _eps_h2omol_unc = 0.0
        _eps_co3_unc = 0.0

        if use_composition_epsilon:
            epsilon_result = calculate_epsilon(
                comp_with_h2o, fe3_over_fetotal=fe3_ratio,
                mole_fractions=density_result["mole_fractions"],
            )
            _eps_oh = epsilon_result["epsilon_H2Ot_3550"]
            _eps_h2omol = epsilon_result["epsilon_H2Om_1635"]
            _eps_co3 = epsilon_result["epsilon_CO3"]
            _eps_oh_unc = epsilon_result["sigma_epsilon_H2Ot_3550"]
            _eps_h2omol_unc = epsilon_result["sigma_epsilon_H2Om_1635"]
            _eps_co3_unc = epsilon_result["sigma_epsilon_CO3"]

        # Compute concentrations
        conc_result = calculate_concentration(
            oh_absorbance=h2o_result["peak_height"],
            h2omol_1630_coeff=co3_result["h2o_1630_coeff"],
            co3_coeff=co3_result["co3_absorbance"],
            density_kgm3=density_result["density_kgm3"],
            thickness_um=_thickness_um,
            epsilon_oh=_eps_oh,
            epsilon_h2omol=_eps_h2omol,
            epsilon_co3=_eps_co3,
            oh_absorbance_unc=h2o_result.get("peak_height_std", 0.0),
            co3_coeff_unc=co3_result.get("co3_total_unc", 0.0),
            thickness_unc_um=thickness_result["stdev_um"],
            epsilon_oh_unc=_eps_oh_unc,
            epsilon_h2omol_unc=_eps_h2omol_unc,
            epsilon_co3_unc=_eps_co3_unc,
            co3_residual_rms=co3_result.get("residual_rms", 0.0),
            co3_snr=co3_result.get("co3_snr", float("inf")),
        )

        new_h2o = conc_result["total_h2o_wt_pct"]
        if new_h2o < 0:
            new_h2o = 0.0

        # Check convergence
        if abs(new_h2o - h2o_estimate) < CONV_THRESHOLD:
            h2o_estimate = new_h2o
            break
        h2o_estimate = new_h2o

    # Store iteration info
    conc_result["density_h2o_iterations"] = _iter + 1
    conc_result["density_h2o_converged"] = (
        _iter + 1 < MAX_ITER or
        abs(new_h2o - h2o_estimate) < CONV_THRESHOLD
    )

    result = {
        "sample_name": sample_name,
        "thickness": thickness_result,
        "density": density_result,
        "refractive_index": ri_result,
        "h2o_peak": {
            "peak_height": h2o_result["peak_height"],
            "peak_wavenumber": h2o_result["peak_wavenumber"],
            "peak_height_std": h2o_result.get("peak_height_std", 0.0),
            "peak_height_range": h2o_result.get("peak_height_range", None),
            "figure_path": h2o_result["figure_path"],
        },
        "co3_fit": {
            "coefficients": co3_result["coefficients"],
            "r_squared": co3_result["r_squared"],
            "co3_absorbance": co3_result["co3_absorbance"],
            "co3_absorbance_unc": co3_result.get("co3_absorbance_unc", 0.0),
            "co3_shift_unc": co3_result.get("co3_shift_unc", 0.0),
            "co3_total_unc": co3_result.get("co3_total_unc", 0.0),
            "h2o_1630_coeff": co3_result["h2o_1630_coeff"],
            "residual_sum_sq": co3_result["residual_sum_sq"],
            "quality_flag": co3_result.get("quality_flag", ""),
            "co3_snr": co3_result.get("co3_snr", 0.0),
            "shift_1430": co3_result.get("shift_1430"),
            "shift_1515": co3_result.get("shift_1515"),
            "fringe_detected": co3_result.get("fringe_detected", False),
            "fringe_amplitude": co3_result.get("fringe_amplitude", 0.0),
            "fringe_period": co3_result.get("fringe_period", 0.0),
            "fringe_ratio": co3_result.get("fringe_ratio", 0.0),
            "fringe_corrected": co3_result.get("fringe_corrected", False),
            "fringe_period_est": co3_result.get("fringe_period_est", 0.0),
            "figure_path": co3_result["figure_path"],
        },
        "concentration": conc_result,
    }
    if epsilon_result is not None:
        result["epsilon"] = epsilon_result
    return result


# ---------------------------------------------------------------------------
# Batch Processing — process_batch
# ---------------------------------------------------------------------------

def process_batch(
    samples: List[Dict],
    output_csv: Optional[Union[str, Path]] = None,
    output_excel: Optional[Union[str, Path]] = None,
    save_figures: bool = True,
    output_dir: Optional[Union[str, Path]] = None,
) -> List[Dict]:
    """Process multiple FTIR samples and optionally write results to file.

    Parameters
    ----------
    samples : list of dict
        Each dict must contain:
            "spectrum_csv"           – path to spectrum CSV
            "composition"            – oxide wt% dict
            "thickness_measurements" – list of (wn_high, wn_low, n_fringes)
        Optional keys:
            "fe3_ratio"    – Fe3+/Fe_total (default 0.15)
            "epsilon_oh"   – absorptivity for OH (default 62.3)
            "epsilon_h2omol" – absorptivity for H2Omol (default 25)
            "epsilon_co3"  – absorptivity for CO3 (default 375)
    output_csv : str or Path or None
        Path for output CSV summary file.
    output_excel : str or Path or None
        Path for output Excel file (requires openpyxl).
    save_figures : bool
        Whether to save per-sample diagnostic figures.
    output_dir : str or Path or None
        Directory for all outputs.

    Returns
    -------
    list of dict – one result dict per sample from process_sample.
    """
    results = []

    for i, sample_info in enumerate(samples, 1):
        name = Path(sample_info["spectrum_csv"]).stem
        print(f"  [{i}/{len(samples)}] Processing {name}...")

        try:
            res = process_sample(
                spectrum_csv=sample_info["spectrum_csv"],
                composition_dict=sample_info["composition"],
                thickness_measurements=sample_info["thickness_measurements"],
                fe3_ratio=sample_info.get("fe3_ratio", 0.15),
                epsilon_oh=sample_info.get("epsilon_oh", 62.3),
                epsilon_h2omol=sample_info.get("epsilon_h2omol", 25.0),
                epsilon_co3=sample_info.get("epsilon_co3", 375.0),
                save_figures=save_figures,
                output_dir=output_dir,
            )
            results.append(res)
        except Exception as e:
            print(f"    ERROR: {e}")
            results.append({"sample_name": name, "error": str(e)})

    # --- Write summary CSV ---
    if output_csv is not None:
        _write_summary_csv(results, output_csv)

    # --- Write summary Excel ---
    if output_excel is not None:
        _write_summary_excel(results, output_excel)

    # --- Print summary table ---
    _print_summary_table(results)

    return results


def _write_summary_csv(results: List[Dict], output_csv: Union[str, Path]):
    """Write batch results to a CSV file."""
    import csv

    headers = [
        "Sample", "Thickness_um", "Thickness_StDev_um",
        "Density_gcc", "Density_kgm3", "Refractive_Index",
        "OH_Abs_3570", "OH_Peak_WN",
        "CO3_Abs", "CO3_Unc", "H2O_1630_Coeff", "CO3_R2", "CO3_Quality",
        "Total_H2O_wt_pct", "H2O_Unc_wt_pct",
        "H2Omol_1630_wt_pct", "CO2_ppm", "CO2_Unc_ppm",
    ]

    with open(output_csv, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)

        for res in results:
            if "error" in res:
                writer.writerow([res["sample_name"], "ERROR: " + res["error"]])
                continue

            writer.writerow([
                res["sample_name"],
                f"{res['thickness']['average_um']:.4f}",
                f"{res['thickness']['stdev_um']:.4f}",
                f"{res['density']['density_gcc']:.6f}",
                f"{res['density']['density_kgm3']:.2f}",
                f"{res['refractive_index']['n']:.6f}",
                f"{res['h2o_peak']['peak_height']:.6f}",
                f"{res['h2o_peak']['peak_wavenumber']:.1f}",
                f"{res['co3_fit']['co3_absorbance']:.6f}",
                f"{res['co3_fit'].get('co3_total_unc', 0.0):.6f}",
                f"{res['co3_fit']['h2o_1630_coeff']:.6f}",
                f"{res['co3_fit']['r_squared']:.6f}",
                res['co3_fit'].get('quality_flag', ''),
                f"{res['concentration']['total_h2o_wt_pct']:.4f}",
                f"{res['concentration'].get('total_h2o_unc_wt_pct', 0.0):.4f}",
                f"{res['concentration']['h2omol_1630_wt_pct']:.4f}",
                f"{res['concentration']['co2_ppm']:.2f}",
                f"{res['concentration'].get('co2_unc_ppm', 0.0):.2f}",
            ])

    print(f"\n  Summary CSV written to: {output_csv}")


def _write_summary_excel(results: List[Dict], output_excel: Union[str, Path]):
    """Write batch results to an Excel file."""
    try:
        import openpyxl
    except ImportError:
        print("  WARNING: openpyxl not installed, skipping Excel output")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FTIR Results"

    headers = [
        "Sample", "Thickness (um)", "Thickness StDev (um)",
        "Density (g/cc)", "Density (kg/m3)", "Refractive Index",
        "OH Abs (3570)", "OH Peak WN",
        "CO3 Abs", "CO3 Unc", "H2O 1630 Coeff", "CO3 R2", "CO3 Quality",
        "Total H2O (wt%)", "H2O Unc (wt%)",
        "H2Omol 1630 (wt%)", "CO2 (ppm)", "CO2 Unc (ppm)",
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)

    for row_idx, res in enumerate(results, 2):
        if "error" in res:
            ws.cell(row_idx, 1, res["sample_name"])
            ws.cell(row_idx, 2, f"ERROR: {res['error']}")
            continue

        values = [
            res["sample_name"],
            res["thickness"]["average_um"],
            res["thickness"]["stdev_um"],
            res["density"]["density_gcc"],
            res["density"]["density_kgm3"],
            res["refractive_index"]["n"],
            res["h2o_peak"]["peak_height"],
            res["h2o_peak"]["peak_wavenumber"],
            res["co3_fit"]["co3_absorbance"],
            res["co3_fit"].get("co3_total_unc", 0.0),
            res["co3_fit"]["h2o_1630_coeff"],
            res["co3_fit"]["r_squared"],
            res["co3_fit"].get("quality_flag", ""),
            res["concentration"]["total_h2o_wt_pct"],
            res["concentration"].get("total_h2o_unc_wt_pct", 0.0),
            res["concentration"]["h2omol_1630_wt_pct"],
            res["concentration"]["co2_ppm"],
            res["concentration"].get("co2_unc_ppm", 0.0),
        ]

        for col, val in enumerate(values, 1):
            ws.cell(row_idx, col, val)

    wb.save(output_excel)
    wb.close()
    print(f"  Summary Excel written to: {output_excel}")


def _print_summary_table(results: List[Dict]):
    """Print a summary table to stdout."""
    print(f"\n{'Sample':<25s} {'d(um)':>8s} {'rho':>8s} {'OH_abs':>8s} "
          f"{'CO3_abs':>8s} {'H2O%':>8s} {'+-':>6s} {'CO2ppm':>10s} {'+-':>8s} {'R2':>8s}")
    print("-" * 115)

    for res in results:
        if "error" in res:
            print(f"{res['sample_name']:<25s}  ERROR: {res['error']}")
            continue

        h2o_unc = res['concentration'].get('total_h2o_unc_wt_pct', 0.0)
        co2_unc = res['concentration'].get('co2_unc_ppm', 0.0)
        print(f"{res['sample_name']:<25s} "
              f"{res['thickness']['average_um']:>8.2f} "
              f"{res['density']['density_gcc']:>8.4f} "
              f"{res['h2o_peak']['peak_height']:>8.4f} "
              f"{res['co3_fit']['co3_absorbance']:>8.4f} "
              f"{res['concentration']['total_h2o_wt_pct']:>8.4f} "
              f"{h2o_unc:>6.4f} "
              f"{res['concentration']['co2_ppm']:>10.2f} "
              f"{co2_unc:>8.2f} "
              f"{res['co3_fit']['r_squared']:>8.6f}")


# ---------------------------------------------------------------------------
# Configuration File Support — load_config / run_from_config
# ---------------------------------------------------------------------------

def load_config(config_path: Union[str, Path]) -> Dict:
    """Load a sample configuration from a YAML or JSON file.

    Parameters
    ----------
    config_path : str or Path
        Path to a .yaml/.yml or .json configuration file.

    Returns
    -------
    dict : parsed configuration with all keys.
    """
    config_path = Path(config_path)
    if not config_path.exists():
        raise FileNotFoundError(f"Config file not found: {config_path}")

    text = config_path.read_text(encoding="utf-8")

    if config_path.suffix in (".yaml", ".yml"):
        try:
            import yaml
        except ImportError:
            raise ImportError(
                "PyYAML is required to read .yaml config files. "
                "Install it with: pip install pyyaml")
        cfg = yaml.safe_load(text)
    elif config_path.suffix == ".json":
        import json
        cfg = json.loads(text)
    else:
        # Try YAML first, then JSON
        try:
            import yaml
            cfg = yaml.safe_load(text)
        except Exception:
            import json
            cfg = json.loads(text)

    if cfg is None:
        raise ValueError(f"Config file is empty: {config_path}")

    # Resolve relative paths relative to the config file's directory
    cfg_dir = config_path.parent
    for key in ("spectrum", "reflectance_csv", "output_dir"):
        if key in cfg and cfg[key] is not None:
            p = Path(cfg[key])
            if not p.is_absolute():
                cfg[key] = str(cfg_dir / p)

    # Resolve relative paths in thickness_measurements is not needed (numeric)
    return cfg


def run_from_config(config_path: Union[str, Path],
                    print_results: bool = True) -> Dict:
    """Run the full FTIR pipeline from a configuration file.

    Parameters
    ----------
    config_path : str or Path
        Path to a .yaml/.yml or .json configuration file.
    print_results : bool
        If True, print a summary to stdout.

    Returns
    -------
    dict : result from process_sample().
    """
    cfg = load_config(config_path)

    # --- Build process_sample keyword arguments ---
    kwargs = {}

    # Required
    if "spectrum" not in cfg:
        raise ValueError("Config must contain 'spectrum' (path to spectrum file)")
    kwargs["spectrum_csv"] = cfg["spectrum"]

    if "composition" not in cfg:
        raise ValueError("Config must contain 'composition' (oxide wt% dict)")
    kwargs["composition_dict"] = cfg["composition"]

    # Thickness (three modes)
    if "thickness_measurements" in cfg and cfg["thickness_measurements"]:
        kwargs["thickness_measurements"] = [
            tuple(m) for m in cfg["thickness_measurements"]
        ]
    if "thickness_um" in cfg and cfg["thickness_um"] is not None:
        kwargs["thickness_um"] = float(cfg["thickness_um"])
    if "thickness_unc_um" in cfg and cfg["thickness_unc_um"] is not None:
        kwargs["thickness_unc_um"] = float(cfg["thickness_unc_um"])
    if "reflectance_csv" in cfg and cfg["reflectance_csv"] is not None:
        kwargs["reflectance_csv"] = cfg["reflectance_csv"]
    if "olivine_fo" in cfg and cfg["olivine_fo"] is not None:
        kwargs["olivine_fo"] = float(cfg["olivine_fo"])

    # Optional parameters
    _OPT_MAP = {
        "fe3_ratio":               ("fe3_ratio", float),
        "co3_model":               ("co3_model", str),
        "co3_fit_range":           ("co3_fit_range", tuple),
        "epsilon_oh":              ("epsilon_oh", float),
        "epsilon_h2omol":          ("epsilon_h2omol", float),
        "epsilon_co3":             ("epsilon_co3", float),
        "use_composition_epsilon": ("use_composition_epsilon", bool),
        "save_figures":            ("save_figures", bool),
        "output_dir":              ("output_dir", str),
    }
    for cfg_key, (param_name, param_type) in _OPT_MAP.items():
        if cfg_key in cfg and cfg[cfg_key] is not None:
            kwargs[param_name] = param_type(cfg[cfg_key])

    # Run pipeline
    result = process_sample(**kwargs)

    # Print summary
    if print_results:
        name = result.get("sample_name", cfg.get("sample_name", "unknown"))
        t = result["thickness"]
        c = result["concentration"]
        co3 = result["co3_fit"]
        model = cfg.get("co3_model", "fixed")
        print(f"\n{'='*60}")
        print(f"  Sample: {name}")
        print(f"  Model:  {model}")
        print(f"{'='*60}")
        print(f"  Thickness: {t['average_um']:.2f} +/- {t['stdev_um']:.2f} um")
        print(f"  H2O:  {c['total_h2o_wt_pct']:.4f} +/- "
              f"{c['total_h2o_unc_wt_pct']:.4f} wt%")
        print(f"  CO2:  {c['co2_ppm']:.1f} +/- {c['co2_unc_ppm']:.1f} ppm")
        print(f"  CO3 R2: {co3['r_squared']:.4f}  [{co3['quality_flag']}]")
        if co3.get("figure_path"):
            print(f"  Figures saved to: "
                  f"{Path(co3['figure_path']).parent}")
        print()

    return result


def run_from_folder(folder_path: Union[str, Path],
                    print_results: bool = True) -> List[Dict]:
    """Discover and run all config files in a folder.

    Searches for *.yaml, *.yml, and *.json files in the given folder
    (non-recursive) and runs each through run_from_config().

    Parameters
    ----------
    folder_path : str or Path
        Directory containing configuration files.
    print_results : bool
        If True, print a summary for each sample.

    Returns
    -------
    list of dict : results from process_sample() for each config.
    """
    folder = Path(folder_path)
    if not folder.is_dir():
        raise NotADirectoryError(f"Not a directory: {folder}")

    configs = sorted(
        list(folder.glob("*.yaml"))
        + list(folder.glob("*.yml"))
        + list(folder.glob("*.json"))
    )
    if not configs:
        raise FileNotFoundError(
            f"No .yaml/.yml/.json config files found in: {folder}")

    print(f"Found {len(configs)} config file(s) in {folder}:")
    for c in configs:
        print(f"  - {c.name}")
    print()

    results = []
    for cfg_path in configs:
        try:
            r = run_from_config(cfg_path, print_results=print_results)
            results.append(r)
        except Exception as e:
            print(f"ERROR processing {cfg_path.name}: {e}\n")

    return results


# ---------------------------------------------------------------------------
# Validation against Excel template data
# ---------------------------------------------------------------------------

def _validate():
    """Run comprehensive validation tests against Excel template values.

    Tests:
      1. Thickness calculation (Reflectance Method) — 2 samples
      2. Glass density (JT12, JT21) — oxide composition to density
      3. Refractive index (Church & Johnson 1980)
      4. H2O peak baseline correction — synthetic spectra + batch
      5. CO3 carbonate fitting — synthetic (5a) + Excel comparison (5b)
      6. Concentration calculations — 5 sub-tests against Master Table
    """
    print("=" * 70)
    print("VALIDATION: Comparing Python output vs Excel template values")
    print("=" * 70)

    # ------------------------------------------------------------------
    # Test 1: Reflectance Method — Sample IH22-qvlb-1
    # ------------------------------------------------------------------
    print("\n--- Test 1: Thickness (Reflectance Method) ---")
    print("Sample: IH22-qvlb-1, refractive index = 1.683 (olivine)")

    measurements = [
        (2882.53, 2423.82, 4),   # JT21_1R
        (2604.33, 2121.02, 4),   # JT21_2R
        (2511.73, 1898.04, 5),   # JT21_3R
    ]
    result = calculate_thickness_multi(measurements, refractive_index=1.683)

    excel_thicknesses = [25.906436082287282, 24.58782416110985, 24.205098044831267]
    excel_avg = 24.899786096076138
    excel_std = 0.8925401761683345

    for i, (calc, excel) in enumerate(zip(result["thicknesses_um"], excel_thicknesses)):
        match = "OK" if abs(calc - excel) < 1e-6 else "MISMATCH"
        print(f"  Measurement {i+1}: calc={calc:.6f} µm | excel={excel:.6f} µm  [{match}]")

    match_avg = "OK" if abs(result["average_um"] - excel_avg) < 1e-6 else "MISMATCH"
    match_std = "OK" if abs(result["stdev_um"] - excel_std) < 1e-6 else "MISMATCH"
    print(f"  Average:   calc={result['average_um']:.6f} µm | excel={excel_avg:.6f} µm  [{match_avg}]")
    print(f"  StDev:     calc={result['stdev_um']:.6f} µm | excel={excel_std:.6f} µm  [{match_std}]")
    print(f"  Thickness: {result['average_um']:.2f} µm = {result['average_cm']:.6f} cm")

    # ------------------------------------------------------------------
    # Test 1b: Reflectance — Sample IH22-qvlb-2
    # ------------------------------------------------------------------
    print("\n--- Test 1b: Thickness — Sample IH22-qvlb-2 ---")
    measurements_2 = [
        (2149.18, 1528.61, 4),   # JT23e_1R
        (2086.47, 1535.84, 4),   # JT23e_2R
        (2294.66, 1929.86, 3),   # JT23e_3R
    ]
    result_2 = calculate_thickness_multi(measurements_2, refractive_index=1.683)
    for i, t in enumerate(result_2["thicknesses_um"]):
        print(f"  Measurement {i+1}: {t:.6f} µm")
    print(f"  Average: {result_2['average_um']:.6f} µm, StDev: {result_2['stdev_um']:.6f} µm")

    # ------------------------------------------------------------------
    # Test 2: Glass Density — Sample JT12
    # ------------------------------------------------------------------
    print("\n--- Test 2: Glass Density (JT12) ---")
    jt12_composition = {
        "SiO2":  50.2776,
        "TiO2":   2.5814566666666665,
        "Al2O3": 14.783166666666668,
        "FeOT":   5.32,
        "MgO":    5.293666666666667,
        "CaO":   10.968066666666667,
        "Na2O":   2.0836366666666666,
        "K2O":    1.3235733333333333,
        "H2O":    0.686,
    }

    density_result = calculate_density(jt12_composition, fe3_over_fetotal=0.15)

    excel_density = 2.6914527126568286
    excel_molar_vol = 23.210244714769825
    excel_gfw = 62.46927609899606

    diff_density = abs(density_result["density_gcc"] - excel_density)
    print(f"  Density:  calc={density_result['density_gcc']:.10f} g/cc")
    print(f"            excel={excel_density:.10f} g/cc")
    print(f"            diff={diff_density:.2e}  [{'OK' if diff_density < 0.001 else 'MISMATCH'}]")
    print(f"  Density:  {density_result['density_kgm3']:.2f} kg/m³")

    diff_mv = abs(density_result["molar_volume"] - excel_molar_vol)
    print(f"  MolarVol: calc={density_result['molar_volume']:.10f}")
    print(f"            excel={excel_molar_vol:.10f}  diff={diff_mv:.2e}")

    diff_gfw = abs(density_result["gfw"] - excel_gfw)
    print(f"  GFW:      calc={density_result['gfw']:.10f}")
    print(f"            excel={excel_gfw:.10f}  diff={diff_gfw:.2e}")

    # ------------------------------------------------------------------
    # Test 2b: Glass Density — Sample JT21
    # ------------------------------------------------------------------
    print("\n--- Test 2b: Glass Density (JT21) ---")
    jt21_composition = {
        "SiO2":  53.2371,
        "TiO2":   2.97705,
        "Al2O3": 15.0541,
        "FeOT":   7.20853,
        "MgO":    5.99398,
        "CaO":   12.9688,
        "Na2O":   2.30455,
        "K2O":    0.450029,
        "H2O":    0.668,
    }

    d21 = calculate_density(jt21_composition, fe3_over_fetotal=0.15)
    excel_d21 = 2.73986  # from the sheet (rounded display)
    diff = abs(d21["density_gcc"] - excel_d21)
    print(f"  Density:  calc={d21['density_gcc']:.6f} g/cc | excel~{excel_d21} g/cc  [diff={diff:.4f}]")
    print(f"  Density:  {d21['density_kgm3']:.2f} kg/m³")

    # ------------------------------------------------------------------
    # Test 3: Refractive Index
    # ------------------------------------------------------------------
    print("\n--- Test 3: Refractive Index (Index of refraction sheet) ---")
    ri_composition = {
        "SiO2":  55.31,
        "TiO2":   2.48,
        "Al2O3": 14.26,
        "FeOT":   3.58,
        "MgO":    3.98,
        "CaO":    7.38,
        "Na2O":   3.53,
        "K2O":    1.41,
        "H2O":    6.42,
    }

    ri_result = get_refractive_index(ri_composition, fe3_over_fetotal=0.15)

    excel_n = 1.519212222247166
    excel_n1 = 0.5192122222471661

    diff_n = abs(ri_result["n"] - excel_n)
    diff_n1 = abs(ri_result["n_minus_1"] - excel_n1)
    print(f"  n-1:   calc={ri_result['n_minus_1']:.12f}")
    print(f"         excel={excel_n1:.12f}  diff={diff_n1:.2e}  [{'OK' if diff_n1 < 1e-6 else 'MISMATCH'}]")
    print(f"  n:     calc={ri_result['n']:.12f}")
    print(f"         excel={excel_n:.12f}  diff={diff_n:.2e}  [{'OK' if diff_n < 1e-6 else 'MISMATCH'}]")

    print(f"\n  Per-oxide contributions to (n-1):")
    for oxide, val in sorted(ri_result["contributions"].items()):
        print(f"    {oxide:8s}: {val:.6f}")

    # ------------------------------------------------------------------
    # Test 4: H2O Peak Baseline Correction (synthetic data)
    # ------------------------------------------------------------------
    print("\n--- Test 4: H2O Peak Baseline Correction (synthetic spectra) ---")

    import tempfile, shutil

    test_dir = Path(tempfile.mkdtemp(prefix="ftir_test_"))
    try:
        # Generate 3 synthetic spectra with different peak heights
        test_params = [
            {"peak_amplitude": 0.45, "peak_center": 3550.0, "seed": 42,
             "name": "sample_A"},
            {"peak_amplitude": 0.30, "peak_center": 3540.0, "seed": 99,
             "name": "sample_B"},
            {"peak_amplitude": 0.60, "peak_center": 3560.0, "seed": 7,
             "name": "sample_C"},
        ]

        for p in test_params:
            csv_path = test_dir / f"{p['name']}.csv"
            _generate_synthetic_spectrum(
                csv_path,
                peak_amplitude=p["peak_amplitude"],
                peak_center=p["peak_center"],
                seed=p["seed"],
            )

        # Test single-file fit with new baseline (low anchor 2200-2400)
        res = fit_h2o_peak(
            test_dir / "sample_A.csv",
            peak_range=(2200, 3800),
            baseline_low_range=(2200, 2400),
            save_figure=True,
        )

        # With the low anchor at 2200-2400 and high anchor at 3700-3800,
        # the broad Gaussian tail (~0.11 at 3750) raises the high anchor,
        # giving a corrected height of ~0.35.  This is expected — real
        # FTIR spectra have a sharper cutoff above 3700 cm-1.
        print(f"  Single file (sample_A):")
        print(f"    Peak height:     {res['peak_height']:.4f}")
        print(f"    Peak wavenumber: {res['peak_wavenumber']:.1f} cm-1  (expected ~3550)")
        print(f"    Data points:     {len(res['wavenumber'])}")
        print(f"    Figure saved:    {res['figure_path']}")

        height_ok = res["peak_height"] > 0.30  # reasonable for synthetic data
        wn_ok = abs(res["peak_wavenumber"] - 3550.0) < 15.0
        print(f"    Height check:    [{'OK' if height_ok else 'MISMATCH'}]")
        print(f"    Position check:  [{'OK' if wn_ok else 'MISMATCH'}]")

        # Test batch processing
        print(f"\n  Batch processing ({test_dir}):")
        batch_results = batch_fit_h2o(test_dir, save_figures=True)

        assert len(batch_results) == 3, f"Expected 3 results, got {len(batch_results)}"
        print(f"  Batch returned {len(batch_results)} results  [OK]")

        # Verify ordering of peak heights matches input amplitudes
        heights = {r["sample"]: r["peak_height"] for r in batch_results}
        assert heights["sample_C"] > heights["sample_A"] > heights["sample_B"], \
            "Peak height ordering mismatch"
        print(f"  Peak height ordering (C > A > B): [OK]")

        # Copy one figure to project dir for visual inspection
        src_fig = res["figure_path"]
        if src_fig and os.path.exists(src_fig):
            dst_fig = os.path.join(
                os.path.dirname(os.path.abspath(__file__)),
                "test_h2o_baseline.png",
            )
            shutil.copy2(src_fig, dst_fig)
            print(f"\n  Figure copied to: {dst_fig}")

    finally:
        shutil.rmtree(test_dir, ignore_errors=True)

    # ------------------------------------------------------------------
    # Test 5: CO₃²⁻ Carbonate Fitting (Module 4)
    # ------------------------------------------------------------------
    print("\n--- Test 5: CO3 Carbonate Fitting ---")

    test_dir = Path(tempfile.mkdtemp(prefix="ftir_co3_test_"))
    try:
        # Generate synthetic spectrum with known coefficients
        known_coeffs = {
            "offset": 0.23,
            "bkg": 0.40,
            "h2o_1630": 0.004,
            "co3": 0.15,
            "slope": -0.0005,
        }
        csv_path = test_dir / "synth_co3.csv"
        _generate_synthetic_co3_spectrum(
            csv_path,
            co3_scale=known_coeffs["co3"],
            bkg_scale=known_coeffs["bkg"],
            h2o_1630_scale=known_coeffs["h2o_1630"],
            offset=known_coeffs["offset"],
            slope=known_coeffs["slope"],
            noise_level=0.001,
            seed=42,
        )

        # Use thickness_cm = 0.01 (= 100 um) so normalization factor = 1
        # and fitted coefficients match the known input coefficients directly.
        result = fit_carbonate(
            csv_path,
            thickness_cm=0.01,
            save_figure=True,
        )

        print(f"  Fitted coefficients:")
        for name, val in result["coefficients"].items():
            known = known_coeffs.get(
                {"offset": "offset", "bkg": "bkg", "c1630": "h2o_1630",
                 "cCO3": "co3", "st": "slope"}[name], None
            )
            if known is not None:
                diff = abs(val - known)
                ok = diff < 0.01
                print(f"    {name:8s}: fitted={val:>10.6f}  known={known:>10.6f}  diff={diff:.2e}  [{'OK' if ok else 'CHECK'}]")
            else:
                print(f"    {name:8s}: fitted={val:>10.6f}")

        print(f"\n  R-squared:        {result['r_squared']:.8f}")
        r2_ok = result["r_squared"] > 0.99
        print(f"  R^2 > 0.99:       [{'OK' if r2_ok else 'MISMATCH'}]")

        print(f"  CO3 peak @1515:   {result['co3_peak_1515']:.6f}")
        print(f"  CO3 peak @1430:   {result['co3_peak_1430']:.6f}")
        print(f"  CO3 absorbance:   {result['co3_absorbance']:.6f}")
        print(f"  Residual SS:      {result['residual_sum_sq']:.6e}")
        print(f"  Figure saved:     {result['figure_path']}")

        # Copy figure for inspection
        if result["figure_path"] and os.path.exists(result["figure_path"]):
            dst = os.path.join(
                os.path.dirname(os.path.abspath(__file__)),
                "test_co3_fit.png",
            )
            shutil.copy2(result["figure_path"], dst)
            print(f"  Figure copied to: {dst}")

        # --- Test 5b: Validate against actual Excel data ---
        print("\n  --- Test 5b: Validate against CO2_bestfit.xls values ---")
        excel_path = Path(__file__).parent / "excel_templates" / "CO2_bestfit.xls"
        if excel_path.exists():
            import xlrd

            wbx = xlrd.open_workbook(str(excel_path), formatting_info=True)
            wsx = wbx.sheet_by_name("CO32-calc")

            # Extract D (raw sample spectrum, col 3) to CSV.
            # Column B (col 1) is already normalized to 100 um; column D is raw.
            # Our fit_carbonate will normalize internally, so we feed it raw data.
            excel_csv = test_dir / "excel_sample.csv"
            with open(excel_csv, "w") as ef:
                ef.write("wavenumber,absorbance\n")
                for row in range(1, 260):
                    wn_val = wsx.cell(row, 0).value
                    d_val = wsx.cell(row, 3).value  # column D = raw spectrum
                    if isinstance(wn_val, float) and isinstance(d_val, float):
                        ef.write(f"{wn_val:.4f},{d_val:.10f}\n")

            thickness_excel = wsx.cell(1, 2).value  # C2 = 0.00301

            res_excel = fit_carbonate(
                excel_csv,
                thickness_cm=thickness_excel,
                save_figure=True,
                figure_path=test_dir / "excel_co3_fit.png",
            )

            # Compare with Excel's fitted coefficients
            excel_coeffs = {
                "offset": wsx.cell(1, 9).value,   # J2
                "bkg": wsx.cell(2, 9).value,       # J3
                "c1630": wsx.cell(3, 9).value,     # J4
                "cCO3": wsx.cell(4, 9).value,      # J5
                "st": wsx.cell(5, 9).value,         # J6
            }
            excel_r2 = wsx.cell(13, 14).value      # O14
            excel_ss = wsx.cell(1, 14).value        # O2

            print(f"  Coefficient comparison (Python vs Excel):")
            for name in ["offset", "bkg", "c1630", "cCO3", "st"]:
                py_val = res_excel["coefficients"][name]
                ex_val = excel_coeffs[name]
                diff = abs(py_val - ex_val)
                pct = abs(diff / ex_val * 100) if ex_val != 0 else 0
                print(f"    {name:8s}: py={py_val:>12.8f}  excel={ex_val:>12.8f}  diff={diff:.2e} ({pct:.2f}%)")

            print(f"\n  R^2:   py={res_excel['r_squared']:.8f}  excel={excel_r2:.8f}  diff={abs(res_excel['r_squared'] - excel_r2):.2e}")
            print(f"  SS(E): py={res_excel['residual_sum_sq']:.8f}  excel={excel_ss:.8f}  diff={abs(res_excel['residual_sum_sq'] - excel_ss):.2e}")

            # Copy figure
            src = test_dir / "excel_co3_fit.png"
            if src.exists():
                dst = os.path.join(
                    os.path.dirname(os.path.abspath(__file__)),
                    "test_co3_fit_excel.png",
                )
                shutil.copy2(str(src), dst)
                print(f"  Figure copied to: {dst}")

            wbx.release_resources()
        else:
            print("  CO2_bestfit.xls not found — skipping Excel validation")

    finally:
        shutil.rmtree(test_dir, ignore_errors=True)

    # ------------------------------------------------------------------
    # Test 6: Concentration Calculations (Module 5 vs Master Table)
    # ------------------------------------------------------------------
    print("\n--- Test 6: Concentration Calculations (Master Table) ---")

    # Test 6a: Total H2O wt% (Q formula)
    # Row 5: IH24-qvbp-2-1
    conc_test_a = calculate_concentration(
        oh_absorbance=0.4085,           # D5
        h2omol_1630_coeff=0.0,          # E5 (not used for Q)
        co3_coeff=0.0,                  # F5 (not used for Q)
        density_kgm3=2691.4527126568287, # O5
        thickness_um=19.564839427527513, # K5
    )
    excel_q = 2.2438600119843324
    diff_q = abs(conc_test_a["total_h2o_wt_pct"] - excel_q)
    print(f"  6a Total H2O (IH24-qvbp-2-1 row 5):")
    print(f"     calc={conc_test_a['total_h2o_wt_pct']:.10f}  excel={excel_q:.10f}  diff={diff_q:.2e}  [{'OK' if diff_q < 1e-6 else 'MISMATCH'}]")

    # Row 43: IH24-qvbp-2-7 (different density & thickness)
    conc_test_b = calculate_concentration(
        oh_absorbance=0.1413,
        h2omol_1630_coeff=0.0,
        co3_coeff=0.0,
        density_kgm3=2794.0449,
        thickness_um=52.1194,
    )
    excel_q2 = 0.28065721
    diff_q2 = abs(conc_test_b["total_h2o_wt_pct"] - excel_q2)
    print(f"  6b Total H2O (IH24-qvbp-2-7 row 43):")
    print(f"     calc={conc_test_b['total_h2o_wt_pct']:.8f}  excel={excel_q2:.8f}  diff={diff_q2:.2e}  [{'OK' if diff_q2 < 1e-4 else 'MISMATCH'}]")

    # Test 6c: H2Omol 1630 wt% (R formula)
    # Row 112: JT49 Run 1
    conc_test_c = calculate_concentration(
        oh_absorbance=0.42,              # D (not tested here)
        h2omol_1630_coeff=0.013323711245520278,  # E112
        co3_coeff=0.0,
        density_kgm3=2891.5876712220834, # O112
        thickness_um=64.6875,            # K112 (needed for Q, not R)
    )
    excel_r = 0.03321265739700762
    diff_r = abs(conc_test_c["h2omol_1630_wt_pct"] - excel_r)
    print(f"  6c H2Omol 1630 (JT49 Run 1 row 112):")
    print(f"     calc={conc_test_c['h2omol_1630_wt_pct']:.12f}  excel={excel_r:.12f}  diff={diff_r:.2e}  [{'OK' if diff_r < 1e-6 else 'MISMATCH'}]")

    # Test 6d: CO2 ppm (V formula)
    # Row 112: JT49 Run 1
    conc_test_d = calculate_concentration(
        oh_absorbance=0.42,
        h2omol_1630_coeff=0.0,
        co3_coeff=0.065169,              # F112
        density_kgm3=2891.5876712220834, # O112
        thickness_um=64.6875,
    )
    excel_v = 264.4404
    diff_v = abs(conc_test_d["co2_ppm"] - excel_v)
    print(f"  6d CO2 ppm (JT49 Run 1 row 112):")
    print(f"     calc={conc_test_d['co2_ppm']:.4f}  excel={excel_v:.4f}  diff={diff_v:.2e}  [{'OK' if diff_v < 0.01 else 'MISMATCH'}]")

    # Test 6e: CO2 ppm — second sample
    # Row 123: JT60 Run 1
    conc_test_e = calculate_concentration(
        oh_absorbance=0.1333,
        h2omol_1630_coeff=0.01994374474593677,
        co3_coeff=0.020212,
        density_kgm3=2724.0,
        thickness_um=49.1541,
    )
    excel_v2 = 87.0591
    excel_r2_val = 0.05277331575943915
    diff_v2 = abs(conc_test_e["co2_ppm"] - excel_v2)
    diff_r2 = abs(conc_test_e["h2omol_1630_wt_pct"] - excel_r2_val)
    print(f"  6e CO2 ppm (JT60 Run 1 row 123):")
    print(f"     CO2:  calc={conc_test_e['co2_ppm']:.4f}  excel={excel_v2:.4f}  diff={diff_v2:.2e}  [{'OK' if diff_v2 < 0.01 else 'MISMATCH'}]")
    print(f"     R:    calc={conc_test_e['h2omol_1630_wt_pct']:.12f}  excel={excel_r2_val:.12f}  diff={diff_r2:.2e}  [{'OK' if diff_r2 < 1e-6 else 'MISMATCH'}]")

    # ------------------------------------------------------------------
    print("\n" + "=" * 70)
    print("VALIDATION COMPLETE")
    print("=" * 70)


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        # No arguments: run validation tests
        _validate()
    else:
        target = sys.argv[1]
        if target in ("--validate", "-v", "test"):
            _validate()
        elif Path(target).is_dir():
            run_from_folder(target)
        elif Path(target).is_file():
            run_from_config(target)
        else:
            print(f"Error: '{target}' is not a valid file or directory.")
            print()
            print("Usage:")
            print("  python ftir_tools.py                   Run validation tests")
            print("  python ftir_tools.py config.yaml       Process single sample")
            print("  python ftir_tools.py configs/           Process all configs in folder")
            print("  python ftir_tools.py --validate         Run validation tests")
            sys.exit(1)
